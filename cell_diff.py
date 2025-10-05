import os
import sys
import re
import subprocess
import errno
import time
import datetime
import shutil
import openpyxl
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

g_left_file      = ''
g_right_file     = ''
g_out_path       = '.'
g_shape_check    = True
g_diff_lib       = True


if g_shape_check:
    import zipfile
    from lxml import etree



DIFF_TEXT_LENGTH = 110
PRE_TEXT_LENGTH  = 24


class cShapeInfo:
    __slots__ = ['id', 'name', 'text', 'col', 'row', 'geom', 'width', 'height', 'sheet']
    def __init__(self):
        self.id     = ''
        self.name   = ''
        self.text   = ''
        self.col    = ''
        self.row    = ''
        self.geom   = ''
        self.width  = 0
        self.height = 0
        self.sheet  = ''
        return

    def print(self):
        print(f'      ID[{self.id}] : {self.geom}, Cell=({self.row},{self.col}), size=(w:{self.width}mm x h:{self.height}mm) : {self.text}')

    def get_geom_text(self):
        self_col    = openpyxl.utils.get_column_letter(self.col)
        return get_disp_string(f'形状({self.geom}), Cell({self_col}{self.row}), Size({self.width}, {self.height})', DIFF_TEXT_LENGTH)

    def get_diff_text(self):
        return get_disp_string(self.text, DIFF_TEXT_LENGTH)


    def compare(self, another):
        if (self.id != another.id) or (self.sheet != another.sheet):
            print(f'      id or sheet unmatch!')
            return

#       print(f'      Compare [col]    : {self.col} vs {another.col}')
#       print(f'      Compare [row]    : {self.row} vs {another.row}')
#       print(f'      Compare [geom]   : {self.geom} vs {another.geom}')
#       print(f'      Compare [width]  : {self.width} vs {another.width}')
#       print(f'      Compare [height] : {self.height} vs {another.height}')
        #/* 幾何学的な情報が異なる場合 */
        self_col    = openpyxl.utils.get_column_letter(self.col)
        another_col = openpyxl.utils.get_column_letter(another.col)
        if (self.col != another.col) or (self.row != another.row) or (self.geom != another.geom) or (self.width != another.width) or (self.height != another.height):
            val_l = self.get_geom_text()
            val_r = another.get_geom_text()
            pre_text = get_disp_string(f'      図形差 ID[{self.id}]', PRE_TEXT_LENGTH)
            print(f'{pre_text} : [{val_l}] vs [{val_r}]')

        if (self.text != another.text):
            pre_text = get_disp_string(f'      Text差 ID[{self.id}]({self_col}{self.row})', PRE_TEXT_LENGTH)
            print(f'{pre_text} : [{self.get_diff_text()}] vs [{another.get_diff_text()}]')



#/*****************************************************************************/
#/* 全角文字の数をカウント                                                    */
#/*****************************************************************************/
def get_full_width_count_in_text(text):
    count = 0
    for character in text:
        if unicodedata.east_asian_width(character) in 'FWA':
            count += 1

    return count


#/*****************************************************************************/
#/* コマンドライン引数処理                                                    */
#/*****************************************************************************/
def check_command_line_option():
    global g_left_file
    global g_right_file
    global g_out_path

    sys.argv.pop(0)
    for arg in sys.argv:
        if (os.path.isfile(arg)):
            if (g_left_file == ''):
                g_left_file = arg
            elif (g_right_file == ''):
                g_right_file = arg
            else:
                print("Too many args! : %s" % arg)
                exit(0)
        else:
            print("invalid arg : %s" % arg)


    if (g_left_file == '') or (g_right_file == ''):
        print("usage : cell_diff.py [file A] [file B]")
        exit(0)

    g_out_path = os.path.dirname(g_right_file)
    if (g_out_path == ''):
        g_out_path = '.'
    return



#/*****************************************************************************/
#/* 処理開始ログ                                                              */
#/*****************************************************************************/
def log_start():
    global g_out_path

    now = datetime.datetime.now()

    time_stamp = now.strftime('%Y%m%d_%H%M%S')
    log_path = g_out_path + '\\cell_diff_' + time_stamp + '.txt'
    log_file = open(log_path, "w")
    sys.stdout = log_file

    start_time = time.perf_counter()
    now = datetime.datetime.now()
    print("処理開始 : " + str(now))
    print ("----------------------------------------------------------------------------------------------------------------")
    return start_time


#/*****************************************************************************/
#/* 処理終了ログ                                                              */
#/*****************************************************************************/
def log_end(start_time):
    end_time = time.perf_counter()
    now = datetime.datetime.now()
    print ("----------------------------------------------------------------------------------------------------------------")
    print("処理終了 : " + str(now))
    second = int(end_time - start_time)
    msec   = ((end_time - start_time) - second) * 1000
    minute = second / 60
    second = second % 60
    print("  %dmin %dsec %dmsec" % (minute, second, msec))
    return


#/*****************************************************************************/
#/* 表示用文字列取得                                                          */
#/*****************************************************************************/
def get_disp_string(text, width):
#   print("get_disp_string Width  : [%d]" % (width))
#   print("get_disp_string Input  : [%s]" % (text))
    cut_flag = 0
    if (result := re.match(r"([^\n]*)\n", text)):
        text = result.group(1)
        cut_flag = 1

    length     = len(text)
    if (length > width):
        text = text[:(width - 3)]
        cut_flag = 1

    length     = len(text)
    full_count = get_full_width_count_in_text(text)
    over_diff  = (length + full_count + (cut_flag * 3)) - width

    while (over_diff > 0):
        cut_flag   = 1
        text = text[:(length - 1)]
        length     = len(text)
        full_count = get_full_width_count_in_text(text)
        length     = len(text)
        full_count = get_full_width_count_in_text(text)
        over_diff  = (length + full_count + (cut_flag * 3)) - width

    if (cut_flag):
        text = text + '...'

    text = text.ljust(width - full_count)
#   print("get_disp_string Output : [%s]" % (text))
    return text


NONE_DATA_TEXT   = get_disp_string('None', DIFF_TEXT_LENGTH)



#/*****************************************************************************/
#/* セル位置が有効範囲内かの判定（0以下の判定はしません）                     */
#/*****************************************************************************/
def is_out_of_bounds(max_row, max_col, row, col):
    if row > max_row:
#       print("Out of Bounds! max_row : [%d] row : [%d]" % (max_row, row))
        return True

    if col > max_col:
#       print("Out of Bounds! max_col : [%d] col : [%d]" % (max_col, col))
        return True

#   print("In Bounds! max[%d, %d],  row, col : [%d, %d]" % (max_row, max_col, row, col))
    return False


#/*****************************************************************************/
#/* 表示用のテキスト比較                                                      */
#/*****************************************************************************/
def get_diff_text(text_l, text_r):
    lines_l = text_l.split("\n")
    lines_r = text_r.split("\n")

    least_index = min([len(lines_l), len(lines_r)])
    for index in range(0, least_index):
#       print(f'  index:{index}, {lines_l[index]} vs {lines_r[index]}')
        if (lines_l[index] != lines_r[index]):
            val_l = get_disp_string(lines_l[index], DIFF_TEXT_LENGTH)
            val_r = get_disp_string(lines_r[index], DIFF_TEXT_LENGTH)
#           print(f'diff in index:{index}, {val_l} vs {val_r}')
            return f'[{val_l}] vs [{val_r}]'

    if (len(lines_l) > len(lines_r)):
        val_l = get_disp_string(lines_l[len(lines_r)], DIFF_TEXT_LENGTH)
        val_r = get_disp_string('',                        DIFF_TEXT_LENGTH)
    elif (len(lines_l) < len(lines_r)):
        val_l = get_disp_string('',                        DIFF_TEXT_LENGTH)
        val_r = get_disp_string(lines_r[len(lines_l)], DIFF_TEXT_LENGTH)
    else:
        val_l = get_disp_string('',                        DIFF_TEXT_LENGTH)
        val_r = get_disp_string('',                        DIFF_TEXT_LENGTH)

    return f'[{val_l}] vs [{val_r}]'


def read_ws_lines(ws):
    lines = []
    for row in ws.iter_rows(values_only=True):
        # セルを文字列化して連結（Noneは空文字にする）
        line = "\t".join("" if v is None else str(v) for v in row)
        lines.append(line)
    return lines


#/*****************************************************************************/
#/* シートの比較(行の追加、削除を検出)                                        */
#/*****************************************************************************/
def check_lr_sheets_ex(ws_l, ws_r):
    print("  シート比較 : [%s]" % (ws_l.title))
    left_add_rows  = []
    right_add_rows = []

    lines_l = read_ws_lines(ws_l)
    lines_r = read_ws_lines(ws_r)
    matcher = SequenceMatcher(None, lines_l, lines_r)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
#           print(f'equal @ {i1+1}, {i2+1}, {j1+1}, {j2+1}')
            continue
        elif tag == 'replace':
#           print(f'replace @ {i1+1}, {i2+1}, {j1+1}, {j2+1}')
            len_l = i2 - i1
            len_r = j2 - j1
            if (len_l > len_r):
                left_add_rows.append(i2)
            elif (len_l < len_r):
                right_add_rows.append(j2)
            continue
        elif tag == 'delete':
#           print(f'delete @ {i1+1}, {i2+1}, {j1+1}, {j2+1}')
            for i in range(i1, i2):
                left_add_rows.append(i+1)
        elif tag == 'insert':
#           print(f'insert @ {i1+1}, {i2+1}, {j1+1}, {j2+1}')
            for j in range(j1, j2):
                right_add_rows.append(j+1)

    max_row_l = ws_l.max_row
    max_col_l = ws_l.max_column
    max_row_r = ws_r.max_row
    max_col_r = ws_r.max_column

    pre_text  = get_disp_string(f'    セル比較', PRE_TEXT_LENGTH)
    row_col_l = get_disp_string(f'max_row = {max_row_l}, max_col = {max_col_l}', DIFF_TEXT_LENGTH)
    row_col_r = get_disp_string(f'max_row = {max_row_r}, max_col = {max_col_r}', DIFF_TEXT_LENGTH)
    print(f'{pre_text} : [{row_col_l}] vs [{row_col_r}]')

    max_row = max([max_row_l, max_row_r])
    max_col = max([max_col_l, max_col_r])


    left_offset = 0
    right_offset = 0
    for row in range(1, max_row + 1):

        if row - left_offset in left_add_rows:
            right_offset += 1
#           print(f'[{row}]right_offset add : {right_offset}')
            for col in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                val_l = ws_l.cell(row - left_offset, col).value
                val_r = "None"
                pre_text = get_disp_string(f'      行削除({col_letter}{row})', PRE_TEXT_LENGTH)
                diff_text = get_diff_text(str(val_l), str(val_r))
                print(f'{pre_text} : {diff_text}')

        elif row - right_offset in right_add_rows:
            left_offset += 1
#           print(f'[{row}]left_offset add : {left_offset}')
            for col in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                val_l = "None"
                val_r = ws_r.cell(row - right_offset, col).value
                pre_text = get_disp_string(f'      行追加({col_letter}{row})', PRE_TEXT_LENGTH)
                diff_text = get_diff_text(str(val_l), str(val_r))
                print(f'{pre_text} : {diff_text}')

        else:
            for col in range(1, max_col + 1):
                val_l = ws_l.cell(row - left_offset, col).value
                val_r = ws_r.cell(row - right_offset, col).value
                if (val_l != val_r):
                    ob_l  = is_out_of_bounds(max_row_l, max_col_l, row, col)
                    ob_r  = is_out_of_bounds(max_row_r, max_col_r, row, col)

                    diff_text = get_diff_text(str(val_l), str(val_r))

                    col_letter = openpyxl.utils.get_column_letter(col)
                    if (ob_l):
                        pre_text = get_disp_string(f'      右増({col_letter}{row})', PRE_TEXT_LENGTH)
                    elif (ob_r):
                        pre_text = get_disp_string(f'      左増({col_letter}{row})', PRE_TEXT_LENGTH)
                    else:
                        pre_text = get_disp_string(f'      差異({col_letter}{row})', PRE_TEXT_LENGTH)

                    print(f'{pre_text} : {diff_text}')
                else:
#                   print(f'left[{row - left_offset}] == right[{row - right_offset}]')
                    pass

    return



#/*****************************************************************************/
#/* シートの比較                                                              */
#/*****************************************************************************/
def check_lr_sheets(ws_l, ws_r):
    print("  シート比較 : [%s]" % (ws_l.title))

    max_row_l = ws_l.max_row
    max_col_l = ws_l.max_column
    max_row_r = ws_r.max_row
    max_col_r = ws_r.max_column

    diff_list = []
    pre_text  = get_disp_string(f'    セル比較', PRE_TEXT_LENGTH)
    row_col_l = get_disp_string(f'max_row = {max_row_l}, max_col = {max_col_l}', DIFF_TEXT_LENGTH)
    row_col_r = get_disp_string(f'max_row = {max_row_r}, max_col = {max_col_r}', DIFF_TEXT_LENGTH)
    print(f'{pre_text} : [{row_col_l}] vs [{row_col_r}]')
    if (max_row_l == max_row_r) and (max_col_l == max_col_r):
        #/* 行、列の数が一致している場合 */
        for row in range(1, max_row_l + 1):
            for col in range(1, max_col_l + 1):
                val_l = ws_l.cell(row, col).value
                val_r = ws_r.cell(row, col).value
                if (val_l != val_r):
                    diff_text = get_diff_text(str(val_l), str(val_r))
                    col_letter = openpyxl.utils.get_column_letter(col)
                    pre_text = get_disp_string(f'      差異({col_letter}{row})', PRE_TEXT_LENGTH)
                    diff_list.append(f'{pre_text} : {diff_text}')

        if len(diff_list):
            for text in diff_list:
                print(text)
        else:
            pre_text  = get_disp_string(f'      全セル一致', PRE_TEXT_LENGTH)
            print(f'{pre_text} : max_row = {max_row_r}, max_col = {max_col_r}')
    else:
        #/* 行、列の数が一致していない場合 */
        max_row = max([max_row_l, max_row_r])
        max_col = max([max_col_l, max_col_r])

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_l = ws_l.cell(row, col).value
                val_r = ws_r.cell(row, col).value
                if (val_l != val_r):
                    ob_l  = is_out_of_bounds(max_row_l, max_col_l, row, col)
                    ob_r  = is_out_of_bounds(max_row_r, max_col_r, row, col)

                    diff_text = get_diff_text(str(val_l), str(val_r))

                    col_letter = openpyxl.utils.get_column_letter(col)
                    if (ob_l):
                        pre_text = get_disp_string(f'      右増({col_letter}{row})', PRE_TEXT_LENGTH)
                    elif (ob_r):
                        pre_text = get_disp_string(f'      左増({col_letter}{row})', PRE_TEXT_LENGTH)
                    else:
                        pre_text = get_disp_string(f'      差異({col_letter}{row})', PRE_TEXT_LENGTH)

                    print(f'{pre_text} : {diff_text}')


    return


#/*****************************************************************************/
#/* 図形の情報比較                                                            */
#/*****************************************************************************/
def check_lr_shapes(sheet, shape_l, shape_r):
    if sheet in shape_l.keys():
#       print("    find l")
        shape_info_l = shape_l[sheet]
    else:
        shape_info_l = []

    if sheet in shape_r.keys():
#       print("    find r")
        shape_info_r = shape_r[sheet]
    else:
        shape_info_r = []

    pre_text  = get_disp_string(f'    図形比較', PRE_TEXT_LENGTH)
    ids_l = get_disp_string(f' {len(shape_info_l)}図形', DIFF_TEXT_LENGTH)
    ids_r = get_disp_string(f' {len(shape_info_r)}図形', DIFF_TEXT_LENGTH)
    print(f'{pre_text} : [{ids_l}] vs [{ids_r}]')

    for l_info in shape_info_l:
        match = False
#       print(f'    check id:{l_info.id}')
        for r_info in shape_info_r:
            if (l_info.id == r_info.id):
                match = True
                l_info.compare(r_info)
                break

        if not match:
            pre_text = get_disp_string(f'      左増 ID[{l_info.id}]', PRE_TEXT_LENGTH)
            geom_text = l_info.get_geom_text()
            diff_text = l_info.get_diff_text()
            print(f'{pre_text} : [{geom_text}] vs [{NONE_DATA_TEXT}]')
            if (l_info.text != ''):
                print(f'{pre_text} : [{diff_text}] vs [{NONE_DATA_TEXT}]')

    for r_info in shape_info_r:
        match = False
        for l_info in shape_info_l:
            if (l_info.id == r_info.id):
                match = True
                break

        if not match:
            pre_text = get_disp_string(f'      右増 ID[{r_info.id}]', PRE_TEXT_LENGTH)
            geom_text = r_info.get_geom_text()
            diff_text = r_info.get_diff_text()
            print(f'{pre_text} : [{NONE_DATA_TEXT}] vs [{geom_text}]')
            if (r_info.text != ''):
                print(f'{pre_text} : [{NONE_DATA_TEXT}] vs [{diff_text}]')


    return


#/*****************************************************************************/
#/* 図形の情報取得                                                            */
#/*****************************************************************************/
def parse_shape_xml(workbook_path):
    shape_list_dic = defaultdict(list)

#   print("  parse shape[%s]" % (workbook_path))
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }

    with zipfile.ZipFile(workbook_path) as z:
        # workbook.xmlを読む
        wb_xml = etree.fromstring(z.read("xl/workbook.xml"))
        # workbook.xml.relsを読む
        wb_rels = etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))

        # rId → sheetX.xml のマップを作る
        rid_to_sheetxml = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in wb_rels
            if "worksheet" in rel.attrib["Type"]
        }
#       print(rid_to_sheetxml)

        # シート名 → ファイルパス の対応を取る
        sheet_name_map = {}
        for sheet in wb_xml.xpath("//main:sheet", namespaces=ns):
            name = sheet.attrib["name"]
            rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            path = "xl/" + rid_to_sheetxml[rid]
            sheet_name_map[path] = name

#       for k, v in sheet_name_map.items():
#           print(f"{k} → {v}")


        # 次にSheetとdrawingのxmlの対応を取る
        sheet_map = {}
        for info in z.infolist():
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
#               print(f'filename : {info.filename}')
                sheet_name = info.filename.split("/")[-1].replace(".xml", "")
                # 対応するrelsを見る
                rel_path = "xl/worksheets/_rels/" + sheet_name + ".xml.rels"
                if rel_path in z.namelist():
                    rel_xml = etree.fromstring(z.read(rel_path))
                    for rel in rel_xml:
                        if "drawing" in rel.attrib["Type"]:
                            target = rel.attrib["Target"].split("/")[-1]
                            sheet_map[target] = sheet_name_map[info.filename]
#                           print(f'    sheet_map : {sheet_map[target]} -> {target}')


        # それぞれのdrawing.xmlを解析
        for target, sheet in sheet_map.items():
#           print(f'    parse {sheet} -> {target}')
            xml = etree.fromstring(z.read("xl/drawings/" + target))
            for anchor in xml.xpath("//xdr:twoCellAnchor | //xdr:oneCellAnchor", namespaces=ns):
                shape_info = cShapeInfo()
                shape_info.sheet = sheet

                # ID
                shape_info.id   = anchor.xpath(".//xdr:cNvPr/@id", namespaces=ns)[0]
                # テキスト
                shape_info.text = '\n'.join(anchor.xpath(".//a:t/text()", namespaces=ns))
                # 図形名
                shape_info.name = anchor.xpath(".//xdr:cNvPr/@name", namespaces=ns)
                # セル座標
                shape_info.col  = int(anchor.xpath(".//xdr:from/xdr:col/text()", namespaces=ns)[0]) + 1
                shape_info.row  = int(anchor.xpath(".//xdr:from/xdr:row/text()", namespaces=ns)[0]) + 1
                # 幾何学情報、サイズ
                shape_info.geom = anchor.xpath(".//a:prstGeom/@prst", namespaces=ns)
                size            = anchor.xpath(".//a:xfrm/a:ext", namespaces=ns)
                if size:
                    cx = int(size[0].attrib["cx"])
                    cy = int(size[0].attrib["cy"])
                    shape_info.width  = int(cx / 914400 * 25.4)
                    shape_info.height = int(cy / 914400 * 25.4)

#               shape_info.print()
#               print(f'sheet : {sheet}')
                shape_list_dic[sheet].append(shape_info)

    return shape_list_dic


#/*****************************************************************************/
#/* ブックの比較                                                              */
#/*****************************************************************************/
def check_lr_books():
    global g_left_file
    global g_right_file

    print(f"ブック比較 : [{g_left_file}] vs [{g_right_file}]")
    wb_l = openpyxl.load_workbook(g_left_file,  data_only=True)
    wb_r = openpyxl.load_workbook(g_right_file, data_only=True)

    if g_shape_check:
        shape_l = parse_shape_xml(g_left_file)
        shape_r = parse_shape_xml(g_right_file)


    for ws_l in wb_l.worksheets:
        for ws_r in wb_r.worksheets:
            if (ws_l.title == ws_r.title):
#               print("title : %s" % ws_r.title)
                if g_diff_lib:
                    check_lr_sheets_ex(ws_l, ws_r)
                else:
                    check_lr_sheets(ws_l, ws_r)


                if g_shape_check:
                    check_lr_shapes(ws_r.title, shape_l, shape_r)
                print('')
    return


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    check_command_line_option()
    start_time = log_start()

     
#   print("TEST1:%s" % (get_disp_string("あいうえおかきくけこ", 10)))
#   print("TEST1:%s" % (get_disp_string("あ\nいうえおかきくけこ", 10)))
    check_lr_books()

    log_end(start_time)
    return


if __name__ == "__main__":
    main()
