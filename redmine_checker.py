import os
import sys
import re
import subprocess
import errno
import time
import datetime
import shutil
import openpyxl
import unicodedata


from pathlib  import Path
from redminelib import Redmine


g_opt_user_name       = 'tkubota'
g_opt_pass            = 'ABCD1234'
g_opt_api_key         = ""
g_opt_full_issues     = 0
g_opt_url             = 'http://localhost:3000/'
g_opt_list_attrs      = ['id']
g_opt_out_file        = 'redmine_result_%ymd.xlsx'
g_opt_in_file         = ''
g_opt_target_projects = []
g_opt_include_sub_prj = 1
g_opt_journal_filters = []
g_opt_redmine_version = None
g_opt_setting_file    = ""
g_opt_grouping        = False
g_opt_cf_format_dic   = {}
g_opt_cf_multi_list   = []
g_opt_issue_list_type = 'flat'

g_target_project_list = []
g_current_user        = None
g_current_user_admin  = False
g_user_list           = []
g_cf_type_list        = []
g_status_type_list    = []
g_issue_list          = []
g_time_entry_list     = []
g_version_list        = []
g_project_id_dic      = {}
g_tracker_id_dic      = {}
g_enum_priority_dic   = {}
g_enum_activity_dic   = {}
g_enum_category_dic   = {}
g_filter_limit        = 20
g_base_day            = datetime.date.today()
g_stats_setting_dic   = {}



ATTR_NAME_DIC      = {
                         'id'                : '#',
                         'project'           : 'プロジェクト',
                         'tracker'           : 'トラッカー',
                         'parent'            : '親チケット',
                         'children'          : '子チケット',
                         'status'            : 'ステータス',
                         'subject'           : 'タイトル',
                         'author'            : '作成者',
                         'created_on'        : '作成日',
                         'priority'          : '優先度',
                         'assigned_to'       : '担当者',
                         'updated_on'        : '更新日',
                         'closed_on'         : '完了日',
                         'due_date'          : '期日',
                         'done_ratio'        : '進捗率',
                         'estimated_hours'   : '予定工数',
                         'total_spent_hours' : '作業時間',
                         'fixed_version'     : '対象バージョン',
                         'journals'          : '更新ID',
                     }



re_1st_line         = re.compile(r"^([^\n]+)\n")
re_key_val_disp     = re.compile(r"^\[([0-9]+)\].+")
re_cf_data          = re.compile(r"^cf_([0-9]+)")
re_issue_id         = re.compile(r"^\#*([0-9]+)")



#/*****************************************************************************/
#/* 辞書型のリスト管理クラス                                                  */
#/*****************************************************************************/
class cListDictionary:
    def __init__(self):
        self.dictionary = {}
        return

    def append_key(self, key):
        if (key not in self.dictionary):
            self.dictionary[key] = []
        return

    def append_item(self, key, item):
        if (key in self.dictionary):
            self.dictionary[key].append(item)
        else:
            self.dictionary[key] = [item]
        return

    def append_wo_duplicate(self, key, item):
        if (key in self.dictionary):
            append_wo_duplicate(self.dictionary[key], item)
        else:
            self.dictionary[key] = [item]
        return

    def get_item_list(self, key):
        if (key in self.dictionary):
            return self.dictionary[key]

        return []


#/*****************************************************************************/
#/* 分析処理クラス                                                            */
#/*****************************************************************************/
class cStatisticsData:
    def __init__(self):
        self.title  = ""
        self.target = ""
        self.unit   = "week"
        self.start  = datetime.date(1970,  1,  1)
        self.end    = datetime.date(2099, 12, 31)
        self.keys   = {1:'-',2:'-',3:'-'}
        return

    def get_next_unit_date(self, some_date):
        if (self.unit == 'month'):
            output_date = get_month_start_day(some_date, 1)
        elif (self.unit == 'day'):
            output_date = some_date + datetime.timedelta(days=1)
        else:
            output_date = some_date + datetime.timedelta(days=7)

        return output_date

g_statistics_data = cStatisticsData()


#/*****************************************************************************/
#/* 統計データキーの登録                                                      */
#/*****************************************************************************/
def check_all_stats_settings():
    global g_stats_setting_dic

    for key, value in g_stats_setting_dic.items():
        if (value.keys[1] == '-'):
            continue

        g_stats_keys.append_wo_duplicate(value.target, value.keys[1])

        if (value.keys[2] == '-'):
            continue

        g_stats_keys.append_wo_duplicate(value.target, value.keys[2])

        if (value.keys[3] == '-'):
            continue

        g_stats_keys.append_wo_duplicate(value.target, value.keys[3])

    return


#/*****************************************************************************/
#/* 統計データ出力設定の取得                                                  */
#/*****************************************************************************/
def get_stats_data(number):
    global g_stats_setting_dic

    if (number in g_stats_setting_dic):
        return g_stats_setting_dic[number]

#   print("new stats data! %s" % number)
    g_stats_setting_dic[number] = cStatisticsData()
    return g_stats_setting_dic[number]


#/*****************************************************************************/
#/* タイムスタンプ保持・更新クラス                                            */
#/*****************************************************************************/
class cTimeStamp:
    def __init__(self, base_ts):
        self.timestamp  = base_ts
        return

    def latter_timestamp(self, ts):
        if (self.timestamp < ts):
            self.timestamp = ts
        return

    def former_timestamp(self, ts):
        if (self.timestamp > ts):
            self.timestamp = ts
        return

g_latest_issues_update_ts = cTimeStamp(datetime.datetime(1970, 1, 1, 0, 0))
g_latest_time_entry_ts    = cTimeStamp(datetime.datetime(1970, 1, 1, 0, 0))
g_first_time_entry_date   = cTimeStamp(datetime.date(2100, 1, 1))



#/*****************************************************************************/
#/* バージョン情報クラス                                                      */
#/*****************************************************************************/
class cVersionData:
    def __init__(self, version):
        self.id     = version.id
        self.name   = version.name
        self.issues = []
        return


#/*****************************************************************************/
#/* プロジェクト情報クラス                                                    */
#/*****************************************************************************/
class cProjectData:
    def __init__(self, project):
        self.id           = project.id
        self.name         = project.name
        self.created_on   = project.created_on
        self.versions     = []
        self.active_users = []
        self.included     = 0                   #/* サブプロジェクト込みで指定された場合1とする */
        return


#/*****************************************************************************/
#/* 作業時間クラス                                                            */
#/*****************************************************************************/
class cTimeEntryData:
    def __init__(self, id):
        self.id           = id
        self.created_on   = ""
        self.user         = None
        self.spent_on     = ""
        self.hours        = 0
        self.updated_on   = ""
        self.activity     = ""
        self.project_name = ""
        self.issue_id     = ""
        return

    def set_data_by_time_entry(self, te):
        self.created_on   = te.created_on
        self.user         = get_user_data_by_id(te.user.id)
        self.spent_on     = te.spent_on
        self.hours        = te.hours
        self.updated_on   = te.updated_on
        self.activity     = str(te.activity)
        self.project_name = te.project.name
        self.issue_id     = getattr_ex(te, 'issue', 'id', 0)
        return


#/*****************************************************************************/
#/* チケット更新の詳細データクラス                                            */
#/*****************************************************************************/
class cDetailData:
    def __init__(self):
        self.property   = ""
        self.name       = ""
        self.old_val    = ""
        self.new_val    = ""
        self.filter     = 0
        return

    def set_detail(self, detail):
#       print("set detail  prop:%s, nmme:%s, old:%s, new:%s" % (detail.get('property'), detail.get('name'), enc_dec_str(detail.get('old_value')), enc_dec_str(detail.get('new_value'))))
        self.property   = detail.get('property')
        self.name       = detail.get('name')
        self.old_val    = omit_multi_line_str(enc_dec_str(detail.get('old_value')))
        self.new_val    = omit_multi_line_str(enc_dec_str(detail.get('new_value')))
        return

    def set_str_data(self, prop, name, old, new):
        self.property   = prop
        if (prop == 'attr'):
            self.name       = name
            self.old_val    = self.set_attr_from_disp_value(old)
            self.new_val    = self.set_attr_from_disp_value(new)
        elif (prop == 'cf'):
            self.name       = str(get_key_from_kv_str(name))
            self.old_val    = self.set_cf_from_disp_value(old)
            self.new_val    = self.set_cf_from_disp_value(new)

#       print("set_str_data prop:%s, name:%s, old:%s, new:%s" % (self.property, self.name, self.old_val, self.new_val))
        self.filter     = 1                       #/* Excelから読み込んだデータは無条件にフィルターをパスする */
        return


    #/*****************************************************************************/
    #/* データ保存対象フィルタのチェック                                          */
    #/*****************************************************************************/
    def filter_check(self):
        global g_opt_journal_filters

        if (self.property == 'attr'):
            for filter_attr in g_opt_journal_filters:
#               print("attr filter [%s] [%s]" % (self.name, filter_attr))
                if (self.name == filter_attr):
                    self.filter = 1
                    return 1
        elif (self.property == 'cf'):
            for filter_attr in g_opt_journal_filters:
                if (self.name == filter_attr):
                    self.filter = 1
                    return 1

        return 0


    #/*****************************************************************************/
    #/* 表示データ名取得                                                          */
    #/*****************************************************************************/
    def get_disp_name(self):
        name = self.name

        #/* カスタムフィールドの場合、nameにはカスタムフィールドのIDがstrで入っているので、カスタムフィールドの名前に変換する */
        if (self.property == 'cf'):
            name = "[%s]%s" % (name, get_custom_field_name(int(name)))

        return name

    #/*****************************************************************************/
    #/* 表示データ変換(Attribute)                                                 */
    #/*****************************************************************************/
    def get_disp_value_attr(self, value):
        if (self.name == 'assigned_to_id'):
            user_data = get_user_data_by_id(int(value))
            value = get_key_value_str(user_data.id, user_data.name)
        elif (self.name == 'status_id'):
            value = get_key_value_str(int(value), get_issue_status_name(int(value)))
        elif (self.name == 'tracker_id'):
            value = get_key_value_str(int(value), g_tracker_id_dic.get(int(value), '不明なトラッカー'))
        elif (self.name == 'priority_id'):
            value = get_key_value_str(int(value), g_enum_priority_dic.get(int(value), '不明なプライオリティ'))
        elif (self.name == 'fixed_version_id'):
            value = get_version_str(int(value))
        elif (self.name == 'done_ratio') or (self.name == 'parent'):
            value = int(value)
        elif (self.name == 'total_spent_hours'):
            value = float(value)
        elif (self.name == 'project_id'):
            value = get_key_value_str(int(value), g_project_id_dic.get(int(value), '不明なプロジェクト'))

        return value

    #/*****************************************************************************/
    #/* 表示データから実データへの変換(Attribute)                                 */
    #/*****************************************************************************/
    def set_attr_from_disp_value(self, text):
        value = text
#       print("set_attr_from_disp_value  text : %s" % text)
        if (text == '-'):
            value = None
        elif (self.name == 'assigned_to_id'):
            value = str(get_key_from_kv_str(text))
        elif (self.name == 'status_id'):
            value = str(get_key_from_kv_str(text))
        elif (self.name == 'tracker_id'):
            value = str(get_key_from_kv_str(text))
        elif (self.name == 'priority_id'):
            value = str(get_key_from_kv_str(text))
        elif (self.name == 'fixed_version_id'):
            value = str(get_key_from_kv_str(text))
        elif (self.name == 'done_ratio') or (self.name == 'parent'):
            value = int(text)
        elif (self.name == 'total_spent_hours'):
            value = float(text)
        elif (self.name == 'project_id'):
            value = str(get_key_from_kv_str(text))

        return value

    #/*****************************************************************************/
    #/* 表示データ変換(Custom Field)                                              */
    #/*****************************************************************************/
    def get_disp_value_cf(self, value):
        id         = int(self.name)
        multiple   = get_custom_field_multiple(id)
        format     = get_custom_field_format(id)
        dictionary = get_custom_field_dictionary(id)

        if (format == 'user'):
            user_data = get_user_data_by_id(int(value))
            value = get_key_value_str(user_data.id, user_data.name)
        elif (format == 'enumeration'):
            value = get_key_value_str(int(value), get_dictionary_value(dictionary, value, value))
        elif (format == 'version'):
            value = get_version_str(int(value))

        return value

    #/*****************************************************************************/
    #/* 表示データから実データへの変換(Custom Field)                              */
    #/*****************************************************************************/
    def set_cf_from_disp_value(self, text):
        value = text
        id         = int(self.name)
        format     = get_custom_field_format(id)
        if (format == 'user'):
            if (value == None) or (value == "") or (value == "-"):
                return None

            value = str(get_key_from_kv_str(text))
        elif (format == 'enumeration'):
            value = str(get_key_from_kv_str(text))
        elif (format == 'version'):
            value = str(get_key_from_kv_str(text))

        return value

    #/*****************************************************************************/
    #/* 表示データ変換                                                            */
    #/*****************************************************************************/
    def get_disp_value(self, value):
        global g_tracker_id_dic
        global g_enum_priority_dic

        if (value == None) or (value == ""):
            return "-"

        if (self.property == 'attr'):
            value = self.get_disp_value_attr(value)
        elif (self.property == 'cf'):
            value = self.get_disp_value_cf(value)

        return value

    def get_disp_old_value(self):
        return self.get_disp_value(self.old_val)

    def get_disp_new_value(self):
        return self.get_disp_value(self.new_val)


#/*****************************************************************************/
#/* チケット更新データクラス                                                  */
#/*****************************************************************************/
class cJournalData:
    def __init__(self, id, created_on, user_data):
        self.id               = id
        self.created_on       = created_on
        self.user             = user_data
        self.details          = []
        self.notes            = ""
        self.filter           = 0
        return

    def is_filter_pass(self):
        global g_opt_journal_filters

        if ('notes' in g_opt_journal_filters):
            if (self.notes != ""):
                return 1

        return self.filter


#/*****************************************************************************/
#/* ユーザーデータクラス                                                      */
#/*****************************************************************************/
class cUserData:
    def __init__(self, id, name):
        self.id           = id
        self.name         = name
        self.time_entries = []
        self.hours_sum    = 0
        return

NONE_USER = cUserData(0, "-")


#/*****************************************************************************/
#/* カスタムフィールドデータクラス                                            */
#/*****************************************************************************/
class cCustomFieldData:
    def __init__(self, cf_id, cf_name, cf_value):
        self.id    = cf_id
        self.name  = cf_name

        format     = get_custom_field_format(cf_id)
        multiple   = get_custom_field_multiple(cf_id)
        dictionary = get_custom_field_dictionary(self.id)

        if (format == 'user'):
            #/* ユーザーIDの場合は、ユーザー情報（cUserData）として情報を保持する。複数可の場合はリスト化する */
            if (cf_value == None) or (cf_value == ''):
                self.value = NONE_USER
            else:
                if (multiple):
                    self.value = []
                    for value in cf_value:
                        self.value.append(get_user_data_by_id(int(value)))
                else:
                    self.value = get_user_data_by_id(int(cf_value)) 
        elif (format == 'enumeration'):
            #/* enumerationのIDの場合は、ID＋テキストの形式で保持する。複数可の場合はリスト化する */
            if (multiple):
#               print("multiple enum : %s" % cf_value)
                self.value = []
                for value in cf_value:
                    str_value = get_dictionary_value(dictionary, value, value)
                    self.value.append(get_key_value_str(int(value), str_value))
            else:
                if (cf_value == None) or (cf_value == '') or (cf_value == '0'):
                    self.value = "[0]-"
                else:
                    str_value = get_dictionary_value(dictionary, cf_value, cf_value)
                    self.value = get_key_value_str(int(cf_value), str_value)
        elif (format == 'version'):
            #/* versionのIDの場合は、ID＋テキストの形式で保持する。複数可の場合はリスト化する */
            if (multiple):
#               print("multiple version : %s" % cf_value)
                self.value = []
                for value in cf_value:
                    self.value.append(get_version_str(int(value)))
            else:
                if (cf_value == None) or (cf_value == '') or (cf_value == '-'):
                    self.value = "[0]-"
                else:
                    self.value = get_version_str(int(cf_value))
        else:
            self.value = cf_value

        return


    #/*****************************************************************************/
    #/* 表示データ変換                                                            */
    #/*****************************************************************************/
    def get_disp_value(self):
        format     = get_custom_field_format(self.id)
        multiple   = get_custom_field_multiple(self.id)

#       print("get_disp_value format:%s, multi:%s" % (format, multiple))
        if (format == "user"):
            if (multiple):
                name_list = []
                for value in self.value:
                    name_list.append(get_key_value_str(value.id, value.name))
                return "\n".join(name_list)                                          #/* 表示名だと重複の可能性があるため、[id]nameのフォーマットで出力する */
            else:
                return get_key_value_str(self.value.id, self.value.name)             #/* 表示名だと重複の可能性があるため、[id]nameのフォーマットで出力する */

        if (multiple):
            ret_val = '\n'.join(self.value)
            return ret_val

        if (type(self.value) is list):
            print("Unknown multiple CF! [%s]:[%s]" % (self.name, self.value))
            ret_val = '\n'.join(self.value)
            return ret_val

        return omit_multi_line_str(enc_dec_str(self.value))



#/*****************************************************************************/
#/* カスタムフィールドの型情報クラス                                          */
#/*****************************************************************************/
class cCustomFieldType:
    def __init__(self, id):
        self.id         = id
        self.name       = ""
        self.type       = 'issue'
        self.format     = "unknown"
        self.dictionary = {}
        self.multiple   = False
        return

    def set_data_by_cf(self, cf):
        global g_opt_cf_format_dic
        global g_opt_cf_multi_list

        self.name       = cf.name
        self.type       = getattr(cf, 'customized_type', 'issue')              #/* 一般ユーザー権限だと取得できない */
        self.format     = getattr(cf, 'field_format', 'unknown')               #/* 一般ユーザー権限だと取得できない */
        self.dictionary = {}
        self.multiple   = getattr(cf, 'multiple', False)

        if (self.name in g_opt_cf_multi_list):
            if (self.format == 'unknown'):
                self.multiple = True
                print("CF[%d]%sを複数選択可能として扱います" % (self.id, self.name))
            elif (self.multiple == False):
                print("CF[%d]%sの複数選択可能設定が異なります" % (self.id, self.name))

        if (self.name in g_opt_cf_format_dic):
            format = g_opt_cf_format_dic[self.name]
            if (self.format == 'unknown'):
                self.format = format
                print("CF[%d]%sのformatを%sとして扱います" % (self.id, self.name, self.format))
            elif (self.format != format):
                print("CF[%d]%sのformatが指定された値と異なります [%s] -> [%s]" % (self.id, self.name, self.format, format))

        if (self.format == 'enumeration'):
            if (hasattr(cf, 'possible_values')):
                for value_label in cf.possible_values:
                    self.dictionary[value_label['value']] = value_label['label']
            else:
                print("管理者権限がないため、enumerationの値は取得できません")

#            print(self.dictionary)
        return


#/*****************************************************************************/
#/* チケットステータス情報クラス                                              */
#/*****************************************************************************/
class cIssueStatusType:
    def __init__(self, status):
        self.id         = status.id
        self.name       = status.name
        self.is_closed  = getattr(status, 'is_closed', 0)
        return


#/*****************************************************************************/
#/* チケットデータクラス                                                      */
#/*****************************************************************************/
class cIssueData:
    def __init__(self, id):
        self.id                = id
        self.project           = ""
        self.parent            = 0
        self.children          = []
        self.priority          = ""
        self.tracker           = ""
        self.subject           = ""
        self.status            = ""
        self.author            = ""
        self.assigned_to       = ""
        self.start_date        = ""
        self.created_on        = ""
        self.updated_on        = ""
        self.closed_on         = ""
        self.done_ratio        = ""
        self.due_date          = ""
        self.estimated_hours   = 0
        self.fixed_version     = "-"
        self.total_spent_hours = 0
        self.journals          = []
        self.time_entries      = []
        self.custom_fields     = []

        self.group_start       = 0             #/* Excel出力の際のGroup化 開始行 */
        self.group_end         = 0             #/* Excel出力の際のGroup化 終了行 */
        return

    #/*****************************************************************************/
    #/* 属性値の取得                                                              */
    #/*****************************************************************************/
    def get_attr_value(self, attribute):
        if (attribute.isdigit()):
            for cf in self.custom_fields:
                if (cf.id == int(attribute)):
                    return cf
        else:
            return getattr(self, attribute)


    #/*****************************************************************************/
    #/* 表示用の属性値取得                                                        */
    #/*****************************************************************************/
    def get_disp_attr(self, attribute):
        if (attribute == 'author') or (attribute == 'assigned_to'):
            object = getattr(self, attribute)
            return get_key_value_str(object.id, object.name)             #/* 表示名だと重複の可能性があるため、[id]nameのフォーマットで出力する */
        elif (attribute == 'children'):
            text = "\n".join(map(str, self.children))
            return text
        elif (attribute.isdigit()):
            for cf in self.custom_fields:
                if (cf.id == int(attribute)):
                    return cf.get_disp_value()

            return ""
        else:
            object = getattr(self, attribute)
            return object


    #/*****************************************************************************/
    #/* python-redmineからチケット情報の読み出し                                  */
    #/*****************************************************************************/
    def read_issue_data(self, issue):
        global g_latest_issues_update_ts
        global g_opt_list_attrs
        global g_current_user_admin

        self.project           = issue.project.name
        self.parent            = getattr_ex(issue, 'parent', 'id', 0)

        children               = getattr(issue, 'children', [])
        self.children          = []
        for child in children:
            self.children.append(child.id)

        self.subject           = enc_dec_str(issue.subject)
        self.priority          = issue.priority.name
        self.tracker           = issue.tracker.name
        self.status            = issue.status.name
        self.author            = get_user_data_by_id(issue.author.id)

        assigned_user          = getattr(issue, 'assigned_to', NONE_USER)
        self.assigned_to       = get_user_data_by_id(assigned_user.id)

        self.created_on        = getattr(issue, 'created_on', "-")
        self.updated_on        = getattr(issue, 'updated_on', "-")
        g_latest_issues_update_ts.latter_timestamp(self.updated_on)

        self.closed_on         = getattr(issue, 'closed_on', "-")
        self.start_date        = getattr(issue, 'start_date', "-")
        self.done_ratio        = getattr(issue, 'done_ratio', 0)
        self.due_date          = getattr(issue, 'due_date', "-")
        self.estimated_hours   = getattr(issue, 'estimated_hours', 0)
        fixed_version          = getattr(issue, 'fixed_version', "-")
        if (fixed_version == "-"):
            self.fixed_version = "[0]-"
        else:
            self.fixed_version = get_version_str(fixed_version.id)
            version_data = get_version_info(fixed_version)
            append_wo_duplicate(version_data.issues, self)               #/* バージョン情報に自分（チケット）の情報を紐づける */


        #/* カスタムフィールドの取得 */
        self.custom_fields = []
        if (hasattr(issue, 'custom_fields')):
            for cf in issue.custom_fields:
#               print("cf[%s]:[%s]" % (cf.name, cf.value))
#               print(dir(cf))
                if not (g_current_user_admin):
                    cf_type = get_custom_field_type(cf.id)
                    if (cf_type == None):
                        cf_type = cCustomFieldType(cf.id)                #/* 一般ユーザーの場合、全プロジェクト共通のCustomFieldは事前には取得できないため、未知のCustomFieldが来るかも */
                        cf_type.set_data_by_cf(cf)                       #/* 一般ユーザーの場合、全プロジェクト共通のCustomFieldは事前には取得できないため、未知のCustomFieldが来るかも */
                        print("  [%d][%s]" % (cf_type.id, cf_type.name))
                        g_cf_type_list.append(cf_type)
                    else:
                        multiple = getattr(cf, 'multiple', False)
                        cf_type.multiple = multiple                      #/* 一般ユーザーはチケットから読み出したCFでないと、複数選択可かどうかがわからない */

                cf_data = cCustomFieldData(cf.id, cf.name, cf.value)
                self.custom_fields.append(cf_data)
                if (str(cf_data.id) in g_opt_list_attrs):
                    pass
                else:
                    print("CF[%d] is not in g_opt_list_attrs" % (cf_data.id))

        #/* 更新情報の取得 */
        self.journals = []
        if (hasattr(issue, 'journals')):
            for journal in issue.journals:
                journal_data = cJournalData(journal.id, journal.created_on, get_user_data_by_id(journal.user.id))
                for detail in journal.details:
                    detail_data = cDetailData()
                    detail_data.set_detail(detail)
#                   print("    detail[%s][%s] %s -> %s" % (detail_data.property, detail_data.name, detail_data.old_val, detail_data.new_val))
                    if (detail_data.filter_check()):
                        journal_data.filter = 1

                    journal_data.details.append(detail_data)

                journal_data.notes = omit_multi_line_str(getattr(journal, 'notes', ""))
                self.journals.append(journal_data)

        #/* 作業時間情報の取得 */
        total_spent_hours = 0
        self.time_entries = []
        if (hasattr(issue, 'time_entries')):
            for time_entry in issue.time_entries:
                te_data = find_time_entry(time_entry.id)
                if (te_data != None):
                    self.time_entries.append(te_data)
                    total_spent_hours += te_data.hours

        #/* total_spent_hoursがサポートされない場合は、time_entriesの合計値とする（本来は、子チケットの時間も集計するようだが・・・） */
        if (hasattr(issue, 'total_spent_hours')):
            self.total_spent_hours = issue.total_spent_hours
        else:
            self.total_spent_hours = total_spent_hours

        return


    #/********************************************************************************/
    #/* 指定した日を越えて最初の更新データを取得する                                 */
    #/********************************************************************************/
    def find_detail_after_some_day(self, attr, some_date):
        last_journal = datetime.date(1900, 1, 1)

        #/* CustomFieldかどうかの判定（attrが数字であればCustomField） */
        if (attr.isdigit()):
            check_prop = 'cf'
            check_name = attr
        else:
            #/* attrの場合、nameの変換が必要 */
            check_prop = 'attr'
            if (attr == 'assigned_to') or (attr == 'status') or (attr == 'fixed_version') or (attr == 'tracker') or (attr == 'priority') or (attr == 'project'):
                check_name = attr + '_id'
            else:
                check_name = attr

        for journal in self.journals:
            created_on_date = journal.created_on.date()
            if (created_on_date < last_journal):
                print("Invalid Journal Sequence! last:%s, next:%s" % (last_journal, created_on_date), file=sys.stderr)
                print("Invalid Journal Sequence! last:%s, next:%s" % (last_journal, created_on_date))
                exit(-1)

            if (some_date < created_on_date):
                for detail in journal.details:
                    if (detail.property == check_prop) and (detail.name == check_name):
                        return detail

            last_journal = journal.created_on.date()
        return None


    #/********************************************************************************/
    #/* チケット属性の日にち指定取得(特定属性のみサポート、複数可のCFはサポート不可) */
    #/********************************************************************************/
    def get_attr_at_some_date(self, attr, some_date, empty_value):
        #/* 指定された日がチケット作成よりも前の日だった場合、指定されたempty_valueを返す */
        if (some_date < self.created_on.date()):
            return empty_value

        last_journal = datetime.date(1900, 1, 1)

        detail = self.find_detail_after_some_day(attr, some_date)
        if (detail == None):
            #/* 更新データが見つからない場合は、現在値を返す（当時の値がそのまま現在の値になっている） */
            return self.get_attr_value(attr)

        if (detail.old_val == None) or (detail.old_val == ''):
            #/* old値が空だった場合は、指定されたempty_valueを返す */
            return empty_value

        if (attr == 'assigned_to'):
            #/* 更新データが見つかった場合は、old値を返す */
            return get_user_data_by_id(int(detail.old_val))
        elif (attr == 'status'):
            return get_issue_status_name(int(detail.old_val))
        elif (attr == 'due_date'):
            return detail.old_val
        elif (attr == 'done_ratio'):
            return int(detail.old_val)
        elif (attr == 'fixed_version'):
            return get_version_str(int(detail.old_val))
        elif (attr == 'priority'):
            return g_enum_priority_dic.get(int(detail.old_val), '不明なプライオリティ')
        elif (attr == 'tracker'):
            return g_tracker_id_dic.get(int(detail.old_val), '不明なトラッカー')
        elif (attr == 'project'):
            return g_project_id_dic.get(int(detail.old_val), '不明なプロジェクト')
        elif (attr.isdigit()):
            cf_id = int(attr)
            return cCustomFieldData(cf_id, get_custom_field_name(cf_id), detail.old_val)
        else:
            print("サポートされないattribute（%s）を指定されました" % (attr))

        return


    #/*****************************************************************************/
    #/* チケット情報のログ出力                                                    */
    #/*****************************************************************************/
    def print_issue_data(self):
        print("--------------------------------- Issue ID : %d ---------------------------------" % (self.id))
        print("  Project         : %s" % (self.project))
        print("  Tracker         : %s" % (self.tracker))
        print("  Subject         : %s" % (self.subject))
        print("  Status          : %s" % (self.status))
        print("  Parent          : %s" % (self.parent))
        print("  Children        : %s" % (self.children))
        print("  Priority        : %s" % (self.priority))
        print("  Author          : %s" % (self.author.name))
        print("  AssignedTo      : %s" % (self.assigned_to.name))
        print("  CreatedOn       : %s" % (self.created_on))
        print("  UpdatedOn       : %s" % (self.updated_on))
        print("  ClosedOn        : %s" % (self.closed_on))
        print("  StartDate       : %s" % (self.start_date))
        print("  DueDate         : %s" % (self.due_date))
        print("  DoneRatio       : %s" % (self.done_ratio))
        print("  EstimatedHours  : %s" % (self.estimated_hours))
        print("  FixedVersion    : %s" % (self.fixed_version))
        print("  TotalSpentHours : %s" % (self.total_spent_hours))
        print("  CustomFields    : ")

        for cf_data in self.custom_fields:
            print("[%s][%s]:%s" % (cf_data.id, cf_data.name, cf_data.get_disp_value()))

        for te_data in self.time_entries:
            print("  Time Entry[%s][%s]:%s hours in %s by %s" % (te_data.id, te_data.created_on, te_data.hours, te_data.spent_on, te_data.user.name))

        for journal_data in self.journals:
            print("  Update[%s][%s]:%s" % (journal_data.id, journal_data.created_on, journal_data.user.name))
            for detail_data in journal_data.details:
                old_val = omit_multi_line_str(detail_data.old_val)
                new_val = omit_multi_line_str(detail_data.new_val)
                if (detail_data.filter_check()):
                    print("    Detail(o)[%s][%s] %s -> %s" % (detail_data.property, detail_data.name, old_val, new_val))
                else:
                    print("    Detail(x)[%s][%s] %s -> %s" % (detail_data.property, detail_data.name, old_val, new_val))

        return


#/*****************************************************************************/
#/* 標準出力とエラー出力の両方に出力                                          */
#/*****************************************************************************/
def print_both(text):
    print(text, file=sys.stderr)
    print(text)
    return


#/*****************************************************************************/
#/* 全角文字の数をカウント                                                    */
#/*****************************************************************************/
def get_full_width_count_in_text(text):
    count = 0
    for character in text:
        if unicodedata.east_asian_width(character) in 'FWA':
            count += 1

    return count


#/*****************************************************************************/
#/* 辞書から値を取得                                                          */
#/*****************************************************************************/
def get_dictionary_value(dictionary, key, default_val):
    if (key in dictionary):
        return dictionary[key]

    return default_val


#/*****************************************************************************/
#/* サブプロジェクトに関するフィルタ取得                                      */
#/*****************************************************************************/
def get_subproject_option():
    global g_opt_include_sub_prj

    if (g_opt_include_sub_prj == 0):
        l_subproject = '!*'
    else:
        l_subproject = '*'

    return l_subproject


#/*****************************************************************************/
#/* 月の初日を返す                                                            */
#/*****************************************************************************/
def get_month_start_day(some_day, offset):
    first_day = some_day.replace(day = 1)
    while(offset > 0):
        first_day = (first_day + datetime.timedelta(days=31)).replace(day = 1)
        offset -= 1

    while(offset < 0):
        first_day = (first_day - datetime.timedelta(days=1)).replace(day = 1)
        offset += 1
    
    return first_day


#/*****************************************************************************/
#/* 週の初日（月曜日）を返す                                                  */
#/*****************************************************************************/
def get_weeks_monday(some_day, offset):
    week_day = some_day.weekday()
    monday = some_day - datetime.timedelta(days=week_day - offset * 7)
    return monday


#/*****************************************************************************/
#/* 重複無しのappend処理                                                      */
#/*****************************************************************************/
def append_wo_duplicate(list, item):
    if (item in list):
        return

    list.append(item)
    return


#/*****************************************************************************/
#/* KEYとVALUEの情報を[KEY]VALUE形式のSTRに変換                               */
#/*****************************************************************************/
def get_key_value_str(key, value):
    return "[%d]%s" % (key, value)


#/*****************************************************************************/
#/* [KEY]VALUE形式のSTRからKEYを取得                                          */
#/*****************************************************************************/
def get_key_from_kv_str(text):
    ret_value = 0

    if (result := re_key_val_disp.match(text)):
        ret_value = int(result.group(1))

    return ret_value


#/*****************************************************************************/
#/* セルの値の取得(整数)                                                      */
#/*****************************************************************************/
def get_cell_value_digit(ws, row, col, none_value):
    value = ws.cell(row, col).value

    if (value == None):
        return none_value
    elif (type(value) is str):
        if (value.isdigit()):
            return int(value)
        else:
            return none_value
    elif (type(value) is int):
        return value

    return none_value


#/*****************************************************************************/
#/* セルの値の取得(チケットID)                                                */
#/*****************************************************************************/
def get_cell_value_issue_id(ws, row, col, none_value):
    id_str = get_cell_value_str(ws, row, col, "")

    if (result := re_issue_id.match(id_str)):
        issue_id = int(result.group(1))
    else:
        issue_id = none_value

    return issue_id


#/*****************************************************************************/
#/* セルの値の取得(文字列)                                                    */
#/*****************************************************************************/
def get_cell_value_str(ws, row, col, none_value):
    value = ws.cell(row, col).value

    if (value == None):
        return none_value
    elif (type(value) is str):
        return value
    elif (type(value) is int):
        return str(value)
    elif (type(value) is float):
        return str(value)
    elif (type(value) is datetime.datetime):
        return value

    print("unknown type! : %s" % type(value))
    return none_value


#/*****************************************************************************/
#/* S-JISでエラーとなる文字の排除                                             */
#/*****************************************************************************/
def enc_dec_str(value):
    if (type(value) is str):
        value = value.encode('cp932', 'replace').decode('cp932', 'replace')

    return value


#/*****************************************************************************/
#/* 複数行のテキストを省略して1行テキストに変換                               */
#/*****************************************************************************/
def omit_multi_line_str(value):
    if (type(value) is str):
        if (result := re_1st_line.match(value)):
            value = result.group(1).replace('\r', '') + '...'                                #/* 改行の含まれる値は無視して1行目だけを扱う */

    return value


#/*****************************************************************************/
#/* 孫attrの取得                                                              */
#/*****************************************************************************/
def getattr_ex(target, attr, sub_attr, default):
    if (hasattr(target, attr)):
        attr = getattr(target, attr)
        ret_val = getattr(attr, sub_attr, default)
    else:
        ret_val = default

    return ret_val


#/*****************************************************************************/
#/* プロジェクト情報の検索                                                    */
#/*****************************************************************************/
def get_project_data_by_name(project_name):
    global g_target_project_list

    for project_data in g_target_project_list:
        if (project_data.name == project_name):
            return project_data

    return None


#/*****************************************************************************/
#/* チケット情報の読み出し                                                    */
#/*****************************************************************************/
def get_issue_data(issue):
    global g_issue_list

    for issue_data in g_issue_list:
        if (issue.id == issue_data.id):
            return issue_data

    issue_data = cIssueData(issue.id)
    g_issue_list.append(issue_data)
    return issue_data


#/*****************************************************************************/
#/* チケット情報の読み出し(id指定)                                            */
#/*****************************************************************************/
def get_issue_data_by_id(issue_id):
    global g_issue_list

    for issue_data in g_issue_list:
        if (issue_id == issue_data.id):
            return issue_data

    #/* リストで見つからない場合は、新規に生成してIDのみを設定しておく */
    issue_data = cIssueData(issue_id)
    g_issue_list.append(issue_data)
    return issue_data


#/*****************************************************************************/
#/* カスタムフィールド情報の取得                                              */
#/*****************************************************************************/
def get_custom_field_type(id):
    global g_cf_type_list

    for cf_type in g_cf_type_list:
        if (cf_type.id == id):
            return cf_type

    return None


#/*****************************************************************************/
#/* カスタムフィールドのフォーマット情報の取得                                */
#/*****************************************************************************/
def get_custom_field_format(id):
    global g_cf_type_list

    cf_type = get_custom_field_type(id)

    if (cf_type != None):
        return cf_type.format

    return "unknown"


#/*****************************************************************************/
#/* カスタムフィールドの名称取得                                              */
#/*****************************************************************************/
def get_custom_field_name(id):
    global g_cf_type_list

    cf_type = get_custom_field_type(id)

    if (cf_type != None):
        return cf_type.name

    return ""


#/*****************************************************************************/
#/* カスタムフィールドのフォーマット情報の取得                                */
#/*****************************************************************************/
def get_custom_field_dictionary(id):
    global g_cf_type_list

    cf_type = get_custom_field_type(id)

    if (cf_type != None):
        return cf_type.dictionary

    return {}


#/*****************************************************************************/
#/* カスタムフィールドが複数選択可能かどうか                                  */
#/*****************************************************************************/
def get_custom_field_multiple(id):
    global g_cf_type_list

    cf_type = get_custom_field_type(id)

    if (cf_type != None):
        return cf_type.multiple

    return False


#/*****************************************************************************/
#/* 作業時間情報の検索                                                        */
#/*****************************************************************************/
def find_time_entry(te_id):
    global g_time_entry_list

    for te_data in g_time_entry_list:
        if (te_data.id == te_id):
            return te_data

    return None


#/*****************************************************************************/
#/* ユーザー情報の登録                                                        */
#/*****************************************************************************/
def get_user_data(redmine, user_id, user_name):
    global g_user_list

    if (user_id == None):
        return cUserData(0, "不明なユーザー")

    for user_data in g_user_list:
        if (user_data.id == user_id) and (user_data.name == user_name):
            return user_data

    user_data = cUserData(user_id, user_name)

    print("[%d]:%s" % (user_data.id, user_data.name))
    g_user_list.append(user_data)
    return user_data


#/*****************************************************************************/
#/* ユーザー情報の登録                                                        */
#/*****************************************************************************/
def get_user_data_by_id(user_id):
    global g_user_list

    if (user_id == 0):
        return NONE_USER

    for user_data in g_user_list:
        if (user_data.id == user_id):
            return user_data

    print("missing user ID! [%d]" % (user_id))
    return cUserData(0, "不明なユーザー")


#/*****************************************************************************/
#/* 設定ファイル読み込み処理                                                  */
#/*****************************************************************************/
def read_setting_file(file_path):
    global g_opt_url
    global g_opt_api_key
    global g_opt_target_projects
    global g_opt_out_file
    global g_opt_in_file
    global g_opt_list_attrs
    global g_opt_include_sub_prj
    global g_opt_journal_filters
    global g_opt_redmine_version
    global g_opt_setting_file
    global g_opt_grouping
    global g_statistics_data
    global g_opt_cf_format_dic
    global g_opt_cf_multi_list
    global g_opt_issue_list_type

    g_opt_setting_file = file_path
    f = open(file_path, 'r', encoding='utf-8')
    lines = f.readlines()

    re_opt_url        = re.compile(r"URL\s+: ([^\n]+)")
    re_opt_api_key    = re.compile(r"API KEY\s+: ([^\n]+)")
    re_opt_out_file   = re.compile(r"OUT FILE NAME\s+: ([^\n]+)")
    re_opt_in_file    = re.compile(r"IN  FILE NAME\s+: ([^\n]+)")
    re_opt_tgt_prj    = re.compile(r"TARGET PROJECT\s+: ([^\n]+)")
    re_opt_list_attr  = re.compile(r"ISSUE LIST ATTR\s+: ([^\n]+)")
    re_opt_list_cf    = re.compile(r"ISSUE LIST CF\s+: ([0-9]+)")
    re_opt_sub_prj    = re.compile(r"INCLUDE SUB PRJ\s+: ([^\n]+)")
    re_opt_filter     = re.compile(r"JOURNAL FILTER\s+: ([^\n]+)")
    re_opt_version    = re.compile(r"REDMINE VERSION\s+: ([^\n]+)")
    re_opt_grouping   = re.compile(r"JOURNAL GROUPING\s+: ([^\n]+)")
    re_opt_sta_unit   = re.compile(r"STATISTICS UNIT\s+: (day|week|month)")
    re_opt_sta_start  = re.compile(r"STATISTICS START\s+: ([0-9]+)[\/\-\s]([0-9]+)[\/\-\s]([0-9]+)")
    re_opt_sta_end    = re.compile(r"STATISTICS END\s+: ([0-9]+)[\/\-\s]([0-9]+)[\/\-\s]([0-9]+)")
    re_opt_cf_format  = re.compile(r"CF FORMAT INFO\s+: ([^ ,]+)[ ,]+([^\n]+)")
    re_opt_cf_multi   = re.compile(r"CF MULTIPLE INFO\s+: ([^\n]+)")
    re_opt_issue_list = re.compile(r"ISSER LIST TYPE\s+: ([^\n]+)")

    re_opt_stats_title = re.compile(r"STATS\[([0-9]+)\] TITLE\s+: ([^\n]+)")
    re_opt_stats_tgt   = re.compile(r"STATS\[([0-9]+)\] TARGET\s+: ([^\n]+)")
    re_opt_stats_unit  = re.compile(r"STATS\[([0-9]+)\] UNIT\s+: ([^\n]+)")
    re_opt_stats_start = re.compile(r"STATS\[([0-9]+)\] START\s+: ([0-9]+)[\/\-\s]([0-9]+)[\/\-\s]([0-9]+)")
    re_opt_stats_end   = re.compile(r"STATS\[([0-9]+)\] END\s+: ([0-9]+)[\/\-\s]([0-9]+)[\/\-\s]([0-9]+)")
    re_opt_stats_key   = re.compile(r"STATS\[([0-9]+)\] KEY\[([0-9]+)\]\s+: ([^\n]+)")

    journal_append = 0
    for line in lines:
#       print ("line:%s" % line)
        if (result := re_opt_url.match(line)):
            g_opt_url = result.group(1)
        elif (result := re_opt_api_key.match(line)):
            g_opt_api_key = result.group(1)
        elif (result := re_opt_tgt_prj.match(line)):
            g_opt_target_projects.append(result.group(1))
        elif (result := re_opt_out_file.match(line)):
            g_opt_out_file = result.group(1)
        elif (result := re_opt_in_file.match(line)):
            g_opt_in_file = result.group(1)
        elif (result := re_opt_list_attr.match(line)):
            if (result.group(1) == 'journals'):
                journal_append = 1                        #/* 表示の都合上、更新情報の出力は一番末尾（右側）とする */
            elif (result.group(1) == 'id'):
                pass                                      #/* IDは必ず先頭にあるので、オプション指定されても無視 */
            else:
                g_opt_list_attrs.append(result.group(1))
        elif (result := re_opt_list_cf.match(line)):
            g_opt_list_attrs.append(result.group(1))
        elif (result := re_opt_sub_prj.match(line)):
            g_opt_include_sub_prj = int(result.group(1))
        elif (result := re_opt_filter.match(line)):
            g_opt_journal_filters.append(result.group(1))
        elif (result := re_opt_version.match(line)):
            g_opt_redmine_version = result.group(1)
        elif (result := re_opt_grouping.match(line)):
            if (result.group(1) == 'True'):
                g_opt_grouping = True
            else:
                g_opt_grouping = False
        elif (result := re_opt_sta_unit.match(line)):
            g_statistics_data.unit = result.group(1)
        elif (result := re_opt_sta_start.match(line)):
            g_statistics_data.start = datetime.date(int(result.group(1)), int(result.group(2)), int(result.group(3)))
        elif (result := re_opt_sta_end.match(line)):
            g_statistics_data.end = datetime.date(int(result.group(1)), int(result.group(2)), int(result.group(3)))
        elif (result := re_opt_cf_format.match(line)):
            key = result.group(2)
            val = result.group(1)
            g_opt_cf_format_dic[key] = val
        elif (result := re_opt_cf_multi.match(line)):
            g_opt_cf_multi_list.append(result.group(1))
        elif (result := re_opt_issue_list.match(line)):
            list_type = result.group(1)
            if (list_type == 'tree') or (list_type == 'flat'):
                g_opt_issue_list_type = list_type
            else:
                print("INVALID ISSUE LIST TYPE : %s" % list_type)
        elif (result := re_opt_stats_title.match(line)):
            stats_data        = get_stats_data(int(result.group(1)))
            stats_data.title  = result.group(2)
#           print("STATS[%s] Title   : %s" % (result.group(1), result.group(2)))
        elif (result := re_opt_stats_tgt.match(line)):
            stats_data        = get_stats_data(int(result.group(1)))
            stats_data.target = result.group(2)
#           print("STATS[%s] Target  : %s" % (result.group(1), result.group(2)))
        elif (result := re_opt_stats_start.match(line)):
            stats_data        = get_stats_data(int(result.group(1)))
            stats_data.start  = datetime.date(int(result.group(2)), int(result.group(3)), int(result.group(4)))
#           print("STATS[%s] Start   : %s" % (result.group(1), stats_data.start))
        elif (result := re_opt_stats_end.match(line)):
            stats_data        = get_stats_data(int(result.group(1)))
            stats_data.end    = datetime.date(int(result.group(2)), int(result.group(3)), int(result.group(4)))
#           print("STATS[%s] End     : %s" % (result.group(1), stats_data.end))
        elif (result := re_opt_stats_key.match(line)):
            stats_data               = get_stats_data(int(result.group(1)))
            key_num                  = int(result.group(2))
            stats_data.keys[key_num] = result.group(3)
#           print("STATS[%s] KEY[%s]  : %s" % (result.group(1), result.group(2), result.group(3)))


    if (journal_append > 0):
        g_opt_list_attrs.append('journals')

    f.close()
    return

#/*****************************************************************************/
#/* コマンドライン引数処理                                                    */
#/*****************************************************************************/
def check_command_line_option():
    global g_opt_user_name
    global g_opt_pass
    global g_opt_full_issues
    global g_opt_api_key

    option = ""
    sys.argv.pop(0)
    for arg in sys.argv:
        if (option == "u"):
            g_opt_user_name = arg
            option = ""
        elif (option == "p"):
            g_opt_pass = arg
            option = ""
        elif (option == "k"):
            g_opt_api_key = arg
            option = ""
        elif (option == "s"):
            read_setting_file(arg)
            option = ""
        elif (arg == "-u") or (arg == "--user"):
            option = "u"
        elif (arg == "-p") or (arg == "--pass"):
            option = "p"
        elif (arg == "-k") or (arg == "--key"):
            option = "k"
        elif (arg == "-s") or (arg == "--set_file"):
            option = "s"
        elif (arg == "-f") or (arg == "--full"):
            g_opt_full_issues = 1
        else:
            print("invalid arg : %s" % arg)

    return



#/*****************************************************************************/
#/* 処理開始ログ                                                              */
#/*****************************************************************************/
def log_start():
    global g_opt_out_file
    now = datetime.datetime.now()

    time_stamp = now.strftime('%Y%m%d_%H%M%S')
    log_path = 'redmine_checker_' + time_stamp + '.txt'
    log_file = open(log_path, "w")
    sys.stdout = log_file

    start_time = time.perf_counter()
    now = datetime.datetime.now()
    print("処理開始 : " + str(now))
    print ("----------------------------------------------------------------------------------------------------------------")

    yyyymmdd = now.strftime('%Y%m%d')
    hhmmss   = now.strftime('%H%M%S')
    g_opt_out_file = g_opt_out_file.replace('%ymd', yyyymmdd)
    g_opt_out_file = g_opt_out_file.replace('%hms', hhmmss)
    return start_time


#/*****************************************************************************/
#/* 処理終了ログ                                                              */
#/*****************************************************************************/
def log_end(start_time):
    end_time = time.perf_counter()
    now = datetime.datetime.now()
    print ("----------------------------------------------------------------------------------------------------------------")
    print("処理終了 : " + str(now))
    second = int(end_time - start_time)
    msec   = ((end_time - start_time) - second) * 1000
    minute = second / 60
    second = second % 60
    print("  %dmin %dsec %dmsec" % (minute, second, msec))
    return


#/*****************************************************************************/
#/* バージョン情報の取得                                                      */
#/*****************************************************************************/
def get_version_info(version):
    global g_version_list

    for version_data in g_version_list:
        if (version.id == version_data.id) and (version.name == version_data.name):
            return version_data

    version_data = cVersionData(version)
    g_version_list.append(version_data)
    return


#/*****************************************************************************/
#/* バージョン情報の取得(ID⇒文字列)                                          */
#/*****************************************************************************/
def get_version_str(version_id):
    global g_version_list

    for version_data in g_version_list:
        if (version_id == version_data.id):
            return "[%d]%s" % (version_id, version_data.name)

    return "[0]-"


#/*****************************************************************************/
#/* プロジェクト情報の登録                                                    */
#/*****************************************************************************/
def add_project_data(project, included):
    global g_current_user
    global g_current_user_admin
    global g_target_project_list

    for member_ship in g_current_user.memberships:
        if (project.name == member_ship.project.name):
            break

    if (g_current_user_admin == False) and (project.name != member_ship.project.name):
        print("このユーザー[%s]は[%s]を参照できません" % (g_current_user.id, project.name))
        return

    project_data = cProjectData(project)
    for version in project.versions:
        print("  Version ID[%d] : %s" % (version.id, version.name))
        version_data = get_version_info(version)
        project_data.versions.append(version_data)

    if (get_project_data_by_name(project_data.name) != None):
        print("プロジェクトの指定が重複しています！ : %s" % (project_data.name), file=sys.stderr)
        print("プロジェクトの指定が重複しています！ : %s" % (project_data.name))
        exit(-1)

    project_data.included = included
    g_target_project_list.append(project_data)
    return


#/*****************************************************************************/
#/* ログインとログインユーザーの権限確認                                      */
#/*****************************************************************************/
def login_and_get_current_user():
    global g_opt_api_key
    global g_opt_user_name
    global g_opt_pass
    global g_opt_redmine_version
    global g_opt_url
    global g_current_user
    global g_current_user_admin

    print("--------------------------------- Login %s ---------------------------------" % (g_opt_url))
    try:
        if (g_opt_api_key == ""):
            redmine = Redmine(g_opt_url, username=g_opt_user_name, password=g_opt_pass, version = g_opt_redmine_version)
        else:
            redmine = Redmine(g_opt_url, key=g_opt_api_key, version = g_opt_redmine_version)

        current_user = redmine.user.get('current')
    except Exception as e:
        print('ログインに失敗しました URL[%s]' % g_opt_url)
        print(f'エラー詳細：{e}')
        exit(-1)

    #/* ログインユーザーのIDと権限を確認 */
    g_current_user = current_user
    current_user_name = enc_dec_str(current_user.lastname) + ' ' + enc_dec_str(current_user.firstname)
    if (hasattr(current_user, 'admin')):
        g_current_user_admin = current_user.admin
        if (g_current_user_admin):
            print("管理者ユーザー:[%d][%s]" % (current_user.id, current_user_name))
        else:
            print("一般ユーザー:[%d][%s]" % (current_user.id, current_user_name))
    else:
        print("管理者情報が取得できません（Redmine4.0.0未満）")
        if (hasattr(current_user, 'mail')):
            g_current_user_admin = True
            print("管理者ユーザー:[%d][%s]" % (current_user.id, current_user_name))
        else:
            g_current_user_admin = False
            print("一般ユーザー:[%d][%s]" % (current_user.id, current_user_name))

#   print(dir(current_user))
    return redmine


#/*****************************************************************************/
#/* プロジェクト情報の取得                                                    */
#/*****************************************************************************/
def check_project_info(redmine):
    global g_opt_target_projects
    global g_target_project_list
    global g_project_id_dic
    global g_opt_include_sub_prj

    print_both("--------------------------------- Check Project Datas ---------------------------------")
    projects = redmine.project.all()

    #/* 対象プロジェクトからIDを取得 */
    for project in projects:
        g_project_id_dic[project.id] = project.name
        for target in g_opt_target_projects:
            if (target == project.name):
                print("ID[%d] : %s" % (project.id, target))
                add_project_data(project, 0)
                break

        #/* ここまで来ている時点で対象に指定されていないプロジェクト */
        if (g_opt_include_sub_prj == 1):
            parent_name = getattr_ex(project, 'parent', 'name', "")
            if (parent_name != ""):
                print("ID[%d]%s has parent [%s]" % (project.id, project.name, parent_name))
                if (get_project_data_by_name(parent_name) != None):
                    #/* 親プロジェクトが登録されていたら、子プロジェクトも登録する */
                    add_project_data(project, 1)

    return


#/*****************************************************************************/
#/* ユーザー情報の取得                                                        */
#/*****************************************************************************/
def check_user_info(redmine):
    global g_current_user_admin
    global g_target_project_list

    print_both("--------------------------------- Check User Datas ---------------------------------")
    l_subproject = get_subproject_option()

    get_all_user = g_current_user_admin

    #/* 全ユーザー情報を取得 */
    try:
        user_ids = [u.id for u in redmine.user.all()]
    except Exception as e:
        print(f"redmine.user.all()を取得できません: {e}")
        get_all_user = 0

    if (get_all_user):
        user_ids = [u.id for u in redmine.user.all()]
        try:
            users = redmine.user.all()
            for user in users:
                print(f"{user.id}: {user.login}")
        except Exception as e:
            print(f"エラーが発生しました: {e}")

        users = redmine.user.all()
        for user in users:
            user_name = enc_dec_str(user.lastname) + ' ' + enc_dec_str(user.firstname)
            get_user_data(redmine, user.id, user_name)

        #/* 全グループ情報も取得 */
        groups = redmine.group.all()
        for group in groups:
            get_user_data(redmine, group.id, enc_dec_str(group.name))
            print("[%d]%s as a Group" % (group.id, enc_dec_str(group.name)))
    else:
        for project_data in g_target_project_list:
            project = redmine.project.get(project_data.id)
            for member_ship in project.memberships:
                if (hasattr(member_ship, 'user')):
                    user = member_ship.user
                    get_user_data(redmine, user.id, enc_dec_str(user.name))
                elif (hasattr(member_ship, 'group')):
                    group = redmine.group.get(member_ship.group.id)
                    get_user_data(redmine, group.id, enc_dec_str(group.name))
                    print("[%d]%s as a Group" % (group.id, enc_dec_str(group.name)))
                else:
                    print("No user or target in MemberShip!")


    return


#/*****************************************************************************/
#/* カスタムフィールド情報の取得                                              */
#/*****************************************************************************/
def check_custom_fields(redmine):
    global g_cf_type_list

    print_both("--------------------------------- Check Custom Fields ---------------------------------")
    if (g_current_user_admin):
        fields = redmine.custom_field.all()
        for cf in fields:
            cf_type = cCustomFieldType(cf.id)
            cf_type.set_data_by_cf(cf)
            if (cf_type.type == "issue"):
                print("  [%d][%s]:%s" % (cf_type.id, cf_type.name, cf_type.format))
                g_cf_type_list.append(cf_type)
    else:
        for project_data in g_target_project_list:
            project = redmine.project.get(project_data.id)
#           print(dir(project))
            for cf in project.issue_custom_fields:
#               print(cf)
#               print(dir(cf))
                cf_type = get_custom_field_type(cf.id)
                if (cf_type == None):
                    cf_type = cCustomFieldType(cf.id)
                    cf_type.set_data_by_cf(cf)
                    print("  [%d][%s]" % (cf_type.id, cf_type.name))
                    g_cf_type_list.append(cf_type)

    return


#/*****************************************************************************/
#/* チケットステータス情報の取得                                              */
#/*****************************************************************************/
def check_issue_status(redmine):
    global g_status_type_list

    statuses = redmine.issue_status.all()

    print_both("--------------------------------- Check Issue Status Types ---------------------------------")
    for status in statuses:
        status_type = cIssueStatusType(status)
        print("  [%d][%s] is_closed : %d" % (status_type.id, status_type.name, status_type.is_closed))
        g_status_type_list.append(status_type)

    return

#/*****************************************************************************/
#/* チケットステータス名の取得                                                */
#/*****************************************************************************/
def get_issue_status_name(status_id):
    global g_status_type_list

    for status in g_status_type_list:
        if (status_id == status.id):
            return status.name

    return "-"


#/*****************************************************************************/
#/* トラッカー名の取得                                                        */
#/*****************************************************************************/
def check_tracker_id(redmine):
    global g_tracker_id_dic

    print_both("--------------------------------- Check Tracker Types ---------------------------------")
    trackers = redmine.tracker.all()

    for tracker in trackers:
        g_tracker_id_dic[tracker.id] = tracker.name
        print("  [%d]%s" % (tracker.id, tracker.name))
    return


#/*****************************************************************************/
#/* 選択肢情報の取得                                                          */
#/*****************************************************************************/
def check_enumerations(redmine):
    global g_enum_priority_dic
    global g_enum_activity_dic
    global g_enum_category_dic

    time_entry_activities = redmine.enumeration.filter(resource='time_entry_activities')
    issue_priorities      = redmine.enumeration.filter(resource='issue_priorities')
    document_categories   = redmine.enumeration.filter(resource='document_categories')

    print_both("--------------------------------- Check Enumeration Types ---------------------------------")
    print("[作業分類]")
    for tea in time_entry_activities:
        print("  [%d]%s" % (tea.id, tea.name))
        g_enum_activity_dic[tea.id] = tea.name

    print("[チケットの優先度]")
    for priority in issue_priorities:
        print("  [%d]%s" % (priority.id, priority.name))
        g_enum_priority_dic[priority.id] = priority.name

    print("[文書カテゴリ]")
    for doc_cat in document_categories:
        print("  [%d]%s" % (doc_cat.id, doc_cat.name))
        g_enum_category_dic[doc_cat.id] = doc_cat.name

    return

#/*****************************************************************************/
#/* 結果フォーマット行出力                                                    */
#/*****************************************************************************/
def output_issue_list_format_line(ws):
    global g_opt_list_attrs

    row = 1
    col = 1
    for item in g_opt_list_attrs:
        if (item in ATTR_NAME_DIC):
            ws.cell(row,     col).value = item
            ws.cell(row + 1, col).value = ATTR_NAME_DIC[item]
            if (item == 'journals'):
                ws.cell(row + 1, col + 1).value = '更新日時'
                ws.cell(row + 1, col + 2).value = '更新者'
                ws.cell(row + 1, col + 3).value = 'コメント'
                ws.cell(row + 1, col + 4).value = '詳細'
                ws.cell(row + 1, col + 5).value = '更新値'
                ws.cell(row + 1, col + 6).value = '更新前'
                ws.cell(row + 1, col + 7).value = '更新後'
        else:
            ws.cell(row,     col).value = 'cf_' + item
            ws.cell(row + 1, col).value = get_custom_field_name(int(item))
        col += 1

    return


#/*****************************************************************************/
#/* 結果フォーマット行出力                                                    */
#/*****************************************************************************/
def output_issue_list_line(ws, row, issue_data, level):
    global g_opt_list_attrs
    global g_opt_issue_list_type

    col = 1
    offset = 0
    for item in g_opt_list_attrs:
        if (item != 'journals'):
#           print("disp attr : %s" % item)
            if (item == 'id') and (g_opt_issue_list_type == 'tree'):
                ws.cell(row, col).value = ('#' * level) + str(issue_data.get_disp_attr(item))
            else:
                ws.cell(row, col).value = issue_data.get_disp_attr(item)
        else:
            for journal in issue_data.journals:
                if (journal.is_filter_pass()):
                    ws.cell(row + offset, col    ).value = journal.id
                    ws.cell(row + offset, col + 1).value = journal.created_on
                    ws.cell(row + offset, col + 2).value = get_key_value_str(journal.user.id, journal.user.name)
                    ws.cell(row + offset, col + 3).value = journal.notes
                    detail_count = 0
                    for detail in journal.details:
                        if (detail.filter):
                            ws.cell(row + offset, col + 4).value = detail.property
                            ws.cell(row + offset, col + 5).value = detail.get_disp_name()
                            ws.cell(row + offset, col + 6).value = detail.get_disp_old_value()
                            ws.cell(row + offset, col + 7).value = detail.get_disp_new_value()
                            detail_count += 1
                            offset += 1

                    if (detail_count == 0):
                        offset += 1                       #/* 詳細データがない場合のみ、次の行に進む */

            col += 6

        col += 1

    if (offset > 0):
        offset -= 1          #/* 1行目のjournalはカウントしないため、引いておく(戻り値はExtraの行数) */

    if (offset > 0):
        #/* Extra行がある場合は、Group化の範囲を覚えておく */
        issue_data.group_start = row + 1
        issue_data.group_end   = row + offset

    print("Issue[%d] row[%d] offset[%d]" % (issue_data.id, row, offset))
    return offset


#/*****************************************************************************/
#/* 結果フォーマット木出力                                                    */
#/*****************************************************************************/
def output_issue_list_tree(ws, row, issue_data, level):
    global g_opt_grouping

    start_row = row
    offset = output_issue_list_line(ws, row, issue_data, level)
#   if (offset > 0):
#       ws.row_dimensions.group(row + 1, row + offset, outline_level=2, hidden=g_opt_grouping)
#       print("Issue[%d] Group[%d]-[%d]" % (issue_data.id, row + 1, row + offset))
    row += (1 + offset)

    for child in issue_data.children:
        child_issue = get_issue_data_by_id(child)
        offset = output_issue_list_tree(ws, row, child_issue, level + 1)
        row += offset

#   print("Issue[%d] row[%d] start[%d] offset[%d]" % (issue_data.id, row, start_row, offset))
    return row - start_row


#/*****************************************************************************/
#/* 結果出力                                                                  */
#/*****************************************************************************/
def output_all_issues_list(ws):
    global g_issue_list
    global g_opt_grouping
    global g_opt_issue_list_type

    print_both("--------------------------------- Output Issue List(%s) ---------------------------------" % (g_opt_issue_list_type))
    output_issue_list_format_line(ws)
    row = 3
    if (g_opt_issue_list_type == 'flat'):
        for issue_data in g_issue_list:
            offset = output_issue_list_line(ws, row, issue_data, 0)
            if (offset > 0):
                ws.row_dimensions.group(row + 1, row + offset, outline_level=1, hidden=g_opt_grouping)
#               print("Issue[%d] Group[%d]-[%d]" % (issue_data.id, row + 1, row + offset))
            row += (1 + offset)
    else:
        for issue_data in g_issue_list:
            if (issue_data.parent == 0):
                offset = output_issue_list_tree(ws, row, issue_data, 0)
                ws.row_dimensions.group(row + 1, row + offset - 1, outline_level=1, hidden=True)
#               print("Root Issue[%d] Group[%d]-[%d]" % (issue_data.id, row + 1, row + offset))
                row += offset

        for issue_data in g_issue_list:
            if (issue_data.group_start != 0):
                ws.row_dimensions.group(issue_data.group_start, issue_data.group_end, outline_level=2, hidden=g_opt_grouping)
#               print("Issue2[%d] Group[%d]-[%d]" % (issue_data.id, issue_data.group_start, issue_data.group_end))

    return


#/*****************************************************************************/
#/* ユーザーごとの作業時間出力                                                */
#/*****************************************************************************/
def output_user_time(ws):
    global g_target_project_list
    global g_time_entry_list
    global g_statistics_data
    global g_user_list
    global g_first_time_entry_date
    global g_base_day

    #/* 最初のTimeEntryと指定された開始日を比較して、実際に出力する開始日を決定する */
    if (g_first_time_entry_date.timestamp < g_statistics_data.start):
        start_day = g_statistics_data.start
    else:
        start_day = g_first_time_entry_date.timestamp

    #/* 今日の日付と指定された終了日を比較して、実際に出力する終了日を決定する */
    if (g_base_day > g_statistics_data.end):
        end_day = g_statistics_data.end
    else:
        end_day = g_base_day

    print_both("--------------------------------- Output User Time Entry [%s --- %s] ---------------------------------" % (start_day, end_day))

    active_user_list = []
    for project_data in g_target_project_list:
        active_user_list.extend(project_data.active_users)
        for user_data in project_data.active_users:
            print("Active for [%s]:[%s]" % (project_data.name, user_data.name))

    active_user_set  = set(active_user_list)
    active_user_set = sorted(active_user_set, key=lambda user: user.id)

    row = 1
    col = 1
    if (g_statistics_data.unit == 'week'):
        ws.cell(row, col).value = '週単位（%s ～ %s）' % (start_day, end_day)
        start_day = get_weeks_monday(start_day, 0)
        end_day   = get_weeks_monday(end_day, 0)
    elif (g_statistics_data.unit == 'month'):
        ws.cell(row, col).value = '月単位（%s ～ %s）' % (start_day, end_day)
        start_day = start_day.replace(day = 1)
        end_day   = end_day.replace(day = 1)
    elif (g_statistics_data.unit == 'day'):
        ws.cell(row, col).value = '日単位（%s ～ %s）' % (start_day, end_day)
    else:
        print("Invalid statistics data unit! : %s" % g_statistics_data.unit)
        return

    output_day = start_day

    row = 3
    col = 1
    for user_data in active_user_set:
        ws.cell(row, col).value = user_data.name
        user_data.hours_sum = 0
        row += 1

    ws.cell(row, col).value = '合計'

    col = 2
    while (output_day <= end_day):
        row = 2
        next_unit = g_statistics_data.get_next_unit_date(output_day)
        print("output : %s - %s" % (output_day, next_unit))
        ws.cell(row, col).value = output_day

        total_hours_in_unit = 0
        for user_data in active_user_set:
            row += 1
            time_spent_in_unit = 0
            for time_entry in user_data.time_entries:
                if (output_day <= time_entry.spent_on) and (time_entry.spent_on < next_unit):
                    time_spent_in_unit += time_entry.hours

            ws.cell(row, col).value = time_spent_in_unit
            total_hours_in_unit += time_spent_in_unit
            user_data.hours_sum += time_spent_in_unit

        row += 1
        ws.cell(row, col).value = total_hours_in_unit            #/* unit合計の出力 */

        output_day = next_unit
        col += 1

    #/* 個人合計の出力 */
    row = 2
    ws.cell(row, col).value = '合計'

    total_hours_in_unit = 0
    for user_data in active_user_set:
        row += 1
        ws.cell(row, col).value = user_data.hours_sum
        total_hours_in_unit += user_data.hours_sum

    row += 1
    ws.cell(row, col).value = total_hours_in_unit

    output_day = next_unit
    col += 1

    return


#/*****************************************************************************/
#/* 統計データ出力（チケット）                                                */
#/*****************************************************************************/
def output_issues_stats(ws, stats):
    return


#/*****************************************************************************/
#/* 統計データ出力（作業時間）                                                */
#/*****************************************************************************/
def output_time_entries_stats(ws, stats):
    print_both("--------------------------------- Output Time Entry Stats [%s] ---------------------------------" % (ws.title))

    #/* 最初のTimeEntryと指定された開始日を比較して、実際に出力する開始日を決定する */
    if (g_first_time_entry_date.timestamp < stats.start):
        start_day = stats.start
    else:
        start_day = g_first_time_entry_date.timestamp

    #/* 今日の日付と指定された終了日を比較して、実際に出力する終了日を決定する */
    if (g_base_day > stats.end):
        end_day = stats.end
    else:
        end_day = g_base_day


    return


#/*****************************************************************************/
#/* 統計データ出力                                                            */
#/*****************************************************************************/
def output_all_stats(wb):
    global g_stats_setting_dic

    for key, value in g_stats_setting_dic.items():
        if (value.target == 'time_entry'):
            ws_title = '%s_%s' % (key, value.title)
            output_time_entries_stats(wb.create_sheet(title = ws_title), value)
        elif (value.target == 'issue'):
            ws_title = '%s_%s' % (key, value.title)
            output_issues_stats(wb.create_sheet(title = ws_title), value)
        else:
            print('')

    return


#/*****************************************************************************/
#/* 作業時間出力                                                              */
#/*****************************************************************************/
def output_all_time_entries(ws):
    global g_time_entry_list

    print_both("--------------------------------- Output All Time Entries ---------------------------------")
    row = 1
    col = 1
    ws.cell(row, col).value = '#'
    col += 1
    ws.cell(row, col).value = '作成日時'
    col += 1
    ws.cell(row, col).value = '作業日'
    col += 1
    ws.cell(row, col).value = '作業時間'
    col += 1
    ws.cell(row, col).value = '作業プロジェクト'
    col += 1
    ws.cell(row, col).value = '作業チケット'
    col += 1
    ws.cell(row, col).value = '作業分類'
    col += 1
    ws.cell(row, col).value = '作業者'
    col += 1
    ws.cell(row, col).value = '更新日'
    col += 1

    row = 2
    for te in g_time_entry_list:
        col = 1
        ws.cell(row, col).value = te.id
        col += 1
        ws.cell(row, col).value = te.created_on
        col += 1
        ws.cell(row, col).value = te.spent_on
        col += 1
        ws.cell(row, col).value = te.hours
        col += 1
        ws.cell(row, col).value = te.project_name
        col += 1
        ws.cell(row, col).value = te.issue_id
        col += 1
        ws.cell(row, col).value = te.activity
        col += 1
        ws.cell(row, col).value = get_key_value_str(te.user.id, te.user.name)
        col += 1
        ws.cell(row, col).value = te.updated_on
        col += 1
        row += 1

    return


#/*****************************************************************************/
#/* 設定値の出力                                                              */
#/*****************************************************************************/
def output_settings(ws):
    global g_opt_url
    global g_opt_setting_file
    global g_target_project_list
    global g_opt_redmine_version
    global g_opt_include_sub_prj
    global g_opt_in_file
    global g_opt_list_attrs
    global g_opt_journal_filters
    global g_opt_full_issues
    global g_base_day
    global g_statistics_data
    global g_opt_cf_format_dic
    global g_opt_cf_multi_list
    global g_current_user
    global g_opt_issue_list_type
    global g_stats_setting_dic

    print_both("--------------------------------- Output Settings ---------------------------------")

    row = 1
    ws.cell(row, 1).value = '実行日'
    ws.cell(row, 2).value = g_base_day
    row += 1

    ws.cell(row, 1).value = 'Redmineバージョン指定'
    if (g_opt_redmine_version != None):
        ws.cell(row, 2).value = g_opt_redmine_version
    else:
        ws.cell(row, 2).value = '-'
    row += 1

    ws.cell(row, 1).value = 'URL'
    ws.cell(row, 2).value = g_opt_url
    row += 1

    ws.cell(row, 1).value = '設定ファイル'
    ws.cell(row, 2).value = g_opt_setting_file
    row += 1

    user_data = get_user_data_by_id(g_current_user.id)
    ws.cell(row, 1).value = '実行ユーザー'
    ws.cell(row, 2).value = get_key_value_str(user_data.id, user_data.name)
    row += 1

    ws.cell(row, 1).value = 'ターゲットプロジェクト'
    col = 2
    for project in g_target_project_list:
        ws.cell(row, col).value = project.name
        col += 1
    row += 1

    ws.cell(row, 1).value = 'サブプロジェクト'
    if (g_opt_include_sub_prj):
        ws.cell(row, 2).value = "含む"
    else:
        ws.cell(row, 2).value = "含まない"
    row += 1

    ws.cell(row, 1).value = '入力ファイル'
    ws.cell(row, 2).value = g_opt_in_file
    row += 1

    ws.cell(row, 1).value = 'チケット一覧形式'
    ws.cell(row, 2).value = g_opt_issue_list_type
    row += 1

    ws.cell(row, 1).value = '表示するチケット属性'
    col = 2
    for attr in g_opt_list_attrs:
        ws.cell(row, col).value = attr
        col += 1
    row += 1

    ws.cell(row, 1).value = '更新チェックフィルター'
    col = 2
    for filter in g_opt_journal_filters:
        ws.cell(row, col).value = filter
        col += 1
    row += 1

    ws.cell(row, 1).value = '全チケットデータ取得'
    col = 2
    if (g_opt_full_issues):
        ws.cell(row, col).value = "YES"
    else:
        ws.cell(row, col).value = "NO"
    row += 1

    ws.cell(row, 1).value = '分析単位'
    col = 2
    ws.cell(row, col).value = g_statistics_data.unit
    row += 1

    ws.cell(row, 1).value = '分析期間'
    col = 2
    ws.cell(row, col).value = g_statistics_data.start
    col += 1
    ws.cell(row, col).value = g_statistics_data.end
    row += 1

    ws.cell(row, 1).value = 'カスタムフィールドのフォーマット'
    col = 2
    for key, value in g_opt_cf_format_dic.items():
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    row += 1

    ws.cell(row, 1).value = 'カスタムフィールドの複数選択可'
    col = 2
    for cf_name in g_opt_cf_multi_list:
        ws.cell(row, col    ).value = cf_name
        row += 1

    row += 1

    for key, value in g_stats_setting_dic.items():
        ws.cell(row, 1).value = '統計データ設定[%d]' % (key)
        col = 2
        row += 1
        ws.cell(row, 1      ).value = '    シート名'
        ws.cell(row, col    ).value = '%s_%s' % (key, value.title)
        row += 1
        ws.cell(row, 1      ).value = '    対象'
        ws.cell(row, col    ).value = value.target
        row += 1
        ws.cell(row, 1      ).value = '    単位'
        ws.cell(row, col    ).value = value.unit
        row += 1
        ws.cell(row, 1      ).value = '    開始'
        ws.cell(row, col    ).value = value.start
        row += 1
        ws.cell(row, 1      ).value = '    終了'
        ws.cell(row, col    ).value = value.end
        row += 1
        ws.cell(row, 1      ).value = '    KEY1'
        ws.cell(row, col    ).value = value.keys[1]
        row += 1
        ws.cell(row, 1      ).value = '    KEY2'
        ws.cell(row, col    ).value = value.keys[2]
        row += 1
        ws.cell(row, 1      ).value = '    KEY3'
        ws.cell(row, col    ).value = value.keys[3]
        row += 1

    return

#/*****************************************************************************/
#/* 各種ID一覧出力                                                            */
#/*****************************************************************************/
def output_id_list(ws):
    global g_user_list
    global g_project_id_dic
    global g_tracker_id_dic
    global g_status_type_list
    global g_version_list
    global g_cf_type_list
    global g_enum_priority_dic
    global g_enum_activity_dic
    global g_enum_category_dic

    print_both("--------------------------------- Output ID List ---------------------------------")

    row = 1
    col = 1
    ws.cell(row, col).value = 'ユーザーID'
    col += 2
    ws.cell(row, col).value = 'プロジェクトID'
    col += 2
    ws.cell(row, col).value = 'トラッカーID'
    col += 2
    ws.cell(row, col).value = 'ステータスID'
    col += 2
    ws.cell(row, col).value = 'バージョンID'
    col += 2
    ws.cell(row, col).value = 'カスタムフィールドID'
    col += 4
    ws.cell(row, col).value = '選択項目'
    col += 2

    row = 2
    col = 1
    for user_data in g_user_list:
        ws.cell(row, col    ).value = user_data.id
        ws.cell(row, col + 1).value = user_data.name
        row += 1

    row = 2
    col = 3
    for key, value in sorted(g_project_id_dic.items()):
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    row = 2
    col = 5
    for key, value in g_tracker_id_dic.items():
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    row = 2
    col = 7
    for status_type in g_status_type_list:
        ws.cell(row, col    ).value = status_type.id
        ws.cell(row, col + 1).value = status_type.name
        row += 1

    row = 2
    col = 9
    for version in g_version_list:
        ws.cell(row, col    ).value = version.id
        ws.cell(row, col + 1).value = version.name
        row += 1

    row = 2
    col = 11
    for cf_type in g_cf_type_list:
        ws.cell(row, col    ).value = cf_type.id
        ws.cell(row, col + 1).value = cf_type.name
        ws.cell(row, col + 2).value = cf_type.format
        if (cf_type.multiple):
            ws.cell(row, col + 3).value = '複数選択可'
        else:
            ws.cell(row, col + 3).value = ''
        row += 1

    row = 2
    col = 15
    ws.cell(row, col    ).value = 'プライオリティ'
    row += 1
    for key, value in g_enum_priority_dic.items():
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    ws.cell(row, col    ).value = '活動種別'
    row += 1
    for key, value in g_enum_activity_dic.items():
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    ws.cell(row, col    ).value = '文書カテゴリ'
    row += 1
    for key, value in g_enum_category_dic.items():
        ws.cell(row, col    ).value = key
        ws.cell(row, col + 1).value = value
        row += 1

    return


#/*****************************************************************************/
#/* 結果出力                                                                  */
#/*****************************************************************************/
def output_datas():
    global g_opt_out_file
    global g_issue_list
    global g_time_entry_list
    global g_user_list
    global g_cf_type_list

    g_issue_list          = sorted(g_issue_list,          key=lambda issue: issue.id)
    g_time_entry_list     = sorted(g_time_entry_list,     key=lambda te:    te.id)
    g_user_list           = sorted(g_user_list,           key=lambda ud:    ud.id)
    g_cf_type_list        = sorted(g_cf_type_list,        key=lambda cf:    cf.id)

    print_both("--------------------------------- Output Redmine Datas ---------------------------------")
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "チケット一覧"
    output_all_issues_list(ws)
    output_all_time_entries(wb.create_sheet(title = "作業時間一覧"))
    output_id_list(wb.create_sheet(title = "ID一覧"))
    output_settings(wb.create_sheet(title = "設定値"))
    output_user_time(wb.create_sheet(title = "ユーザー作業時間"))
#   output_all_stats(wb)
    wb.save(g_opt_out_file)
    return


#/*****************************************************************************/
#/* 作業時間の読み出し、登録                                                  */
#/*****************************************************************************/
def read_time_entries(time_entries, user_data, project_data):
    global g_latest_time_entry_ts
    global g_first_time_entry_date

    for time_entry in time_entries:
        te = find_time_entry(time_entry.id)
        if (te == None):
            te = cTimeEntryData(time_entry.id)
            user_data.time_entries.append(te)

        te.set_data_by_time_entry(time_entry)
        g_latest_time_entry_ts.latter_timestamp(te.updated_on)

        print("  [%s][%s] %s, %s h for #%s (%s), updated : %s" % (te.id, te.created_on, te.spent_on, te.hours, te.issue_id, te.activity, te.updated_on))
        g_first_time_entry_date.former_timestamp(te.spent_on)
        append_wo_duplicate(g_time_entry_list, te)

    if (len(time_entries) > 0):
        print("Project[%s] Active[%s]" % ((project_data.name, user_data.name)))
        append_wo_duplicate(project_data.active_users, user_data)

    return


#/*****************************************************************************/
#/* 作業時間チェック                                                          */
#/*****************************************************************************/
def time_entry_check(redmine, is_full_check):
    global g_time_entry_list
    global g_target_project_list
    global g_user_list
    global g_latest_time_entry_ts

    l_subproject = get_subproject_option()
    if (is_full_check):
        print_both("--------------------------------- Time Entries Check Full ---------------------------------")
        for user_data in g_user_list:
            for project_data in g_target_project_list:
                if (project_data.included):
                    continue                         #/* 子プロジェクトは親チケットと一緒にfilterされるので、スキップする */

                time_entries = redmine.time_entry.filter(project_id = project_data.id, subproject_id = l_subproject, user_id = user_data.id)
                read_time_entries(time_entries, user_data, project_data)
    else:
        last_updated = g_latest_time_entry_ts.timestamp.date()

        print_both("--------------------------------- Time Entries Check onwards %s ---------------------------------" % (last_updated))
        for user_data in g_user_list:
            for project_data in g_target_project_list:
                if (project_data.included):
                    continue                         #/* 子プロジェクトは親チケットと一緒にfilterされるので、スキップする */
                time_entries = redmine.time_entry.filter(project_id = project_data.id, subproject_id = l_subproject, user_id = user_data.id, from_date = last_updated)      #/* Time Entry は from_dateでフィルタする */
                read_time_entries(time_entries, user_data, project_data)

    return


#/*****************************************************************************/
#/* チケット全確認                                                            */
#/*****************************************************************************/
def full_issue_check(redmine):
    global g_filter_limit

    print_both("--------------------------------- Full Issue Check ---------------------------------")
    for project_data in g_target_project_list:
        if (project_data.included):
            continue                         #/* 子プロジェクトは親チケットと一緒にfilterされるので、スキップする */

        l_subproject = get_subproject_option()
        filter_offset = 0
        while(1):
            issues = redmine.issue.filter(project_id = project_data.id, subproject_id = l_subproject, status_id = '*', limit = g_filter_limit, offset = filter_offset)
            if (len(issues) == 0):
                break

            print_both("--------------------------------- ProjectID : %d, Filter Offset %d ---------------------------------" % (project_data.id, filter_offset))
            for issue in issues:
                issue_data = get_issue_data(issue)
                issue_data.read_issue_data(issue)
                issue_data.print_issue_data()

            filter_offset += g_filter_limit

    return


#/*****************************************************************************/
#/* チケット確認(新規・更新)                                                  */
#/*****************************************************************************/
def issue_check(redmine):
    global g_filter_limit
    global g_latest_issues_update_ts

    last_updated = g_latest_issues_update_ts.timestamp.date()
    print_both("--------------------------------- Created / Updated Issue Check ---------------------------------")
    for project_data in g_target_project_list:
        if (project_data.included):
            continue                         #/* 子プロジェクトは親チケットと一緒にfilterされるので、スキップする */

        l_subproject = get_subproject_option()
        updated_option = '>=%s' % last_updated
        filter_offset = 0
        while(1):
            issues = redmine.issue.filter(project_id = project_data.id, subproject_id = l_subproject, status_id = '*', limit = g_filter_limit, offset = filter_offset, updated_on = updated_option)
            if (len(issues) == 0):
                break

            print_both("--------------------------------- ProjectID : %d, Updated onwards %s Filter Offset %d ---------------------------------" % (project_data.id, last_updated, filter_offset))
            for issue in issues:
                issue_data = get_issue_data(issue)
                issue_data.read_issue_data(issue)
                issue_data.print_issue_data()

            filter_offset += g_filter_limit

    return


#/*****************************************************************************/
#/* チケット属性読み出し                                                    */
#/*****************************************************************************/
def read_journal_lines(ws, row, issue_data, j_col, d_col):
    offset = 0
    next_id = issue_data.id
    while(1):
        journal_id  = get_cell_value_str(ws, row + offset, j_col,     "")
        detail_prop = get_cell_value_str(ws, row + offset, d_col    , "")
        detail_name = get_cell_value_str(ws, row + offset, d_col + 1, "")
        detail_old  = get_cell_value_str(ws, row + offset, d_col + 2, "")
        detail_new  = get_cell_value_str(ws, row + offset, d_col + 3, "")
#       print("row = %d, jID = %s, prop = %s, next_id = %d" % (row + offset, journal_id, detail_prop, next_id))

        #/* 「チケット番号、更新ID、詳細がすべて空」 OR 「次の行のIDが0以外で、異なるIDであった場合、ループ終了」 */
        if ((journal_id == "") and (detail_prop == "") and (next_id == 0)) or ((next_id != 0) and (next_id != issue_data.id)):
            break

        if (journal_id != ""):
            journal_data = cJournalData(int(journal_id), "", NONE_USER)
            journal_data.created_on = get_cell_value_str(ws, row + offset, j_col + 1, "")

            user_str                = get_cell_value_str(ws, row + offset, j_col + 2, "")
            journal_data.user       = get_user_data_by_id(get_key_from_kv_str(user_str))
            journal_data.notes      = get_cell_value_str(ws, row + offset, j_col + 3, "")
            journal_data.filter     = 1
            issue_data.journals.append(journal_data)

        if (detail_prop != ""):
            detail_data = cDetailData()
            detail_data.set_str_data(detail_prop, detail_name, detail_old, detail_new)
            journal_data.details.append(detail_data)


        offset += 1

        #/* 次の行のチケットIDを取得 */
        next_id = get_cell_value_issue_id(ws, row + offset, 1, 0)


    if (offset > 0):
        offset -= 1

    return offset


#/*****************************************************************************/
#/* チケット属性読み出し                                                    */
#/*****************************************************************************/
def read_attr_value(issue_data, attr, value):
    global g_latest_issues_update_ts

    if (result := re_cf_data.match(attr)):
        #/* カスタムフィールドの場合 */
        if (value == ""):
            return

        cf_id    = int(result.group(1))
        cf_name  = get_custom_field_name(cf_id)
        multiple = get_custom_field_multiple(cf_id)
        format   = get_custom_field_format(cf_id)

        if (format == 'user') or (format == 'version') or (format == 'enumeration'):
            #/* カスタムフィールドでユーザー／バージョン／enum情報の場合 */
            if (multiple):
                cf_val = []
#               print("cf user/version/enum multiple : %s" % value)
                if (value == None) or (value == ""):
                    pass
                else:
                    tmp_vals = value.split("\n")
                    for tmp_val in tmp_vals:
                        tmp_val_int = get_key_from_kv_str(tmp_val)
                        if (tmp_val_int != 0):
                            cf_val.append(str(tmp_val_int))
            else:
#               print("get cf_%d, user : %s" % (cf_id, value))
                cf_val = str(get_key_from_kv_str(value))
        elif (format == 'date'):
            if (type(value) is datetime.datetime):
                cf_val = value.date()
            else:
                cf_val = value
        elif (multiple):
            cf_val = []
            tmp_vals = value.split("\n")
            for tmp_val in tmp_vals:
                cf_val.append(tmp_val)
        else:
            cf_val = value

        cf_data = cCustomFieldData(cf_id, cf_name, cf_val)
        issue_data.custom_fields.append(cf_data)

    elif (attr == 'author') or (attr == 'assigned_to'):
        #/* ユーザー情報の場合 */
        setattr(issue_data, attr, get_user_data_by_id(get_key_from_kv_str(value)))
    elif (attr == 'children'):
        if (value == ""):
            issue_data.children = []
        else:
            tmp_vals = value.split("\n")
            for tmp_val in tmp_vals:
                issue_data.children.append(int(tmp_val))
    elif (attr == 'done_ratio') or (attr == 'parent'):
        setattr(issue_data, attr, int(value))
    elif (attr == 'total_spent_hours') or (attr == 'estimated_hours'):
        if (value == None) or (value == ''):
            setattr(issue_data, attr, None)
        else:
            setattr(issue_data, attr, float(value))
    elif (attr == 'due_date'):
        if (type(value) is datetime.datetime):
            setattr(issue_data, attr, value.date())
        else:
            setattr(issue_data, attr, value)
    elif (attr == 'created_on'):
        setattr(issue_data, attr, value)
    elif (attr == 'updated_on'):
        setattr(issue_data, attr, value)
        g_latest_issues_update_ts.latter_timestamp(value)
    elif (attr != ''):
        setattr(issue_data, attr, value)

    return



#/*****************************************************************************/
#/* 作業時間一覧の読み込み                                                    */
#/*****************************************************************************/
def read_time_entry_list(ws):
    global g_time_entry_list
    global g_latest_time_entry_ts
    print_both("--------------------------------- Read Time Entry List! ---------------------------------")

    row = 1
    col = 1
    id_def = get_cell_value_str(ws, row, col, "")
    if (id_def != '#'):
        print("  このシートは読み込めません！")
        return

    row = 2
    while (ws.cell(row, 1).value != None):
        col = 1

        #/* ID(int) */
        te_id = get_cell_value_digit(ws, row, col, 0)
        col += 1

        #/* 作成日時(datetime) */
        te_created_on = get_cell_value_str(ws, row, col, 0)
        col += 1

        #/* 作業日（date） */
        te_spent_on = get_cell_value_str(ws, row, col, 0)
        te_spent_on = te_spent_on.date()
        col += 1

        #/* 作業時間(float) */
        te_hours = float(get_cell_value_str(ws, row, col, 0))
        col += 1

        #/* プロジェクト(str) */
        te_project = get_cell_value_str(ws, row, col, 0)
        col += 1

        #/* チケット番号(int) */
        te_issue_id = get_cell_value_digit(ws, row, col, 0)
        col += 1

        #/* 活動内容(str) */
        te_activity = get_cell_value_str(ws, row, col, 0)
        col += 1

        #/* ユーザー */
        te_user = get_user_data_by_id(get_key_from_kv_str(ws.cell(row, col).value))
        col += 1

        #/* 更新日 */
        te_updated_on = get_cell_value_str(ws, row, col, 0)
        g_latest_time_entry_ts.latter_timestamp(te_updated_on)
        col += 1

        te = cTimeEntryData(te_id)
        te.created_on   = te_created_on
        te.spent_on     = te_spent_on
        te.hours        = te_hours
        te.user         = te_user
        te.project_name = te_project
        te.issue_id     = te_issue_id
        te.updated_on   = te_updated_on
        te.activity     = te_activity
        print("  [%s][%s] %s, %s h for #%s (%s), updated : %s" % (te.id, te.created_on, te.spent_on, te.hours, te.issue_id, te.activity, te.updated_on))
        te_user.time_entries.append(te)
        g_time_entry_list.append(te)
        project_data = get_project_data_by_name(te.project_name)
        if (project_data == None):
            print("  不明なプロジェクトに対する作業時間です : %s" % (te.project_name))
        else:
            append_wo_duplicate(project_data.active_users, te.user)
        g_first_time_entry_date.former_timestamp(te.spent_on)

        issue_data = get_issue_data_by_id(te.issue_id)
        issue_data.time_entries.append(te)

        row += 1

    return

#/*****************************************************************************/
#/* チケット一覧の読み込み                                                    */
#/*****************************************************************************/
def read_issue_list(ws):
    print_both("--------------------------------- Read Issue List! ---------------------------------")

    row = 1
    col = 1
    id_def = get_cell_value_str(ws, row, col, "")
    if (id_def != 'id'):
        print("  このシートは読み込めません！")
        return

    #/* 更新データの所在を確認する */
    journal_col = 0
    detail_col = 0
    while (ws.cell(row, col).value != None):
        if (ws.cell(row, col).value == 'journals'):
            journal_col = col
            detail_col  = col + 4
            break
        col += 1

    if (journal_col == 0):
        journal_col = col                       #/* 見つからなかった場合は、末尾の列を指しておく(detail_colはゼロのまま) */

#   print("journal_col:%d" % (journal_col))

    row = 3
    while (1):
        col = 1
        issue_id = get_cell_value_issue_id(ws, row, col, 0)
        if (issue_id == 0):
            if (detail_col == 0):
                #/* 更新データの列がない場合は、即終了 */
                break
            else:
                journal_id  = get_cell_value_digit(ws, row, journal_col, 0)
                detail_prop = get_cell_value_str(ws, row, detail_col, "")
                if (journal_id == 0) and (detail_prop == ""):
                    #/* IDもJournalもDetailもなければ終了 */
                    break
        else:
            print("  read issue id[%d]" % issue_id)
            issue_data = get_issue_data_by_id(issue_id)
            for col in range(2, journal_col):
                attr_name = get_cell_value_str(ws, 1, col, "")
#               print("attr:%s, value:%s" % (attr_name, get_cell_value_str(ws, row, col, "")))
                read_attr_value(issue_data, attr_name, get_cell_value_str(ws, row, col, ""))

            #/* 更新データの読み込み */
            if (detail_col > 0):
                offset = read_journal_lines(ws, row, issue_data, journal_col, detail_col)
                row += offset
        row += 1
    return


#/*****************************************************************************/
#/* ID一覧の読み込み                                                          */
#/*****************************************************************************/
def read_id_list(ws):
    global g_current_user_admin
    global g_cf_type_list

    print_both("--------------------------------- Read ID List! ---------------------------------")

    row = 1
    col = 1

    while (col < 100):
        if (ws.cell(row, col).value == 'カスタムフィールドID'):
            break
        col += 1

    if (col >= 100):
        print("カスタムフィールドに関する情報が見つかりませんでした")
        return

    row = 2
    while (1):
        id = ws.cell(row, col).value
        if (id == None):
            break

        name   = ws.cell(row, col + 1).value
        format = ws.cell(row, col + 2).value
        multi  = ws.cell(row, col + 3).value
        if (multi == '複数選択可'):
            multiple = True
        else:
            multiple = False

        cf_type = get_custom_field_type(id)
        if (cf_type == None):
            print("CF[%d][%s]の情報が見つかりません" % (id, name))
            if (g_current_user_admin):
                pass                                                 #/* 管理者権限の場合は全情報を取得しているので、何もしない */
            else:
                cf_type          = cCustomFieldType(id)              #/* 一般権限の場合はこの時点で登録する                     */
                cf_type.name     = name
                cf_type.format   = format
                cf_type.multiple = multiple
                g_cf_type_list.append(cf_type)
        else:
            if (g_current_user_admin):
                #/* 管理者権限の場合は全情報を取得しているので、ログを吐くだけ */
                if (cf_type.name != name):
                    print("CF[%d]の名前が変わっています [%s] -> [%s]" % (id, name, cf_type.name))
                if (cf_type.format != format):
                    print("CF[%d]のフォーマットが変わっています [%s] -> [%s]" % (id, format, cf_type.format))
                if (cf_type.multiple != multiple):
                    print("CF[%d]の複数選択可否が変わっています [%s] -> [%s]" % (id, multiple, cf_type.multiple))
            else:
                #/* 一般権限の場合は名前が変わっている場合は処理中止する */
                if (cf_type.name != name):
                    print("CF[%d]の名前が変わっています [%s] -> [%s]" % (id, name, cf_type.name))
                    exit(-1)

                if (cf_type.format != format):
                    print("CF[%d]のフォーマットが変わっています [%s] -> [%s]" % (id, format, cf_type.format))
                    cf_type.format = format

                if (cf_type.multiple != multiple):
                    print("CF[%d]の複数選択可否が変わっています [%s] -> [%s]" % (id, multiple, cf_type.multiple))
                    cf_type.multiple = multiple

        row += 1

    return


#/*****************************************************************************/
#/* 前回出力ファイルの読み込み                                                */
#/*****************************************************************************/
def read_in_file():
    global g_opt_in_file

    print_both("--------------------------------- Read Input File : %s ---------------------------------" % (g_opt_in_file))
    if (g_opt_in_file == ''):
        return

    wb = openpyxl.load_workbook(g_opt_in_file, data_only=True)

    for ws in wb.worksheets:
        if (ws.title == "チケット一覧"):
            read_issue_list(ws)
        elif (ws.title == "作業時間一覧"):
            read_time_entry_list(ws)
        elif (ws.title == "ID一覧"):
            read_id_list(ws)
        else:
            print("skip ws : %s" % ws.title)

    wb.close()
    return



#/*****************************************************************************/
#/* チケットの過去データ参照（試験用）                                        */
#/*****************************************************************************/
def test_print_old_issue_status(issue_data, some_day):
    user_data     = issue_data.get_attr_at_some_date('assigned_to',   some_day, NONE_USER)
    user_name     = get_key_value_str(user_data.id, user_data.name)
    count         = get_full_width_count_in_text(user_name)
    user_name     = user_name.ljust(16 - count)

    status        = issue_data.get_attr_at_some_date('status',        some_day, '新規')
    count         = get_full_width_count_in_text(status)
    status        = status.ljust(14 - count)

    done_ratio    = issue_data.get_attr_at_some_date('done_ratio',    some_day, 0)
    done_ratio    = str(done_ratio).rjust(3)

    due_date      = issue_data.get_attr_at_some_date('due_date',      some_day, '')
    due_date      = str(due_date).ljust(10)

    project       = issue_data.get_attr_at_some_date('project',       some_day, '')
    count         = get_full_width_count_in_text(project)
    project       = project.ljust(18 - count)

    tracker       = issue_data.get_attr_at_some_date('tracker',       some_day, '')
    count         = get_full_width_count_in_text(tracker)
    tracker       = tracker.ljust(18 - count)

    fixed_version = issue_data.get_attr_at_some_date('fixed_version', some_day, '')
    count         = get_full_width_count_in_text(fixed_version)
    fixed_version = fixed_version.ljust(14 - count)

    print("  [%s] : assigned[%s], status[%s], done_ratio[%s], due_date[%s], project[%s], tracker[%s], fixed[%s]" % (some_day, user_name, status, done_ratio, due_date, project, tracker, fixed_version))
    return


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    global g_opt_full_issues

    check_command_line_option()
    start_time = log_start()

    redmine = login_and_get_current_user()
    check_project_info(redmine)
    check_user_info(redmine)
    check_custom_fields(redmine)
    check_issue_status(redmine)
    check_tracker_id(redmine)
    check_enumerations(redmine)

    if (g_opt_full_issues):
        time_entry_check(redmine, 1)
        full_issue_check(redmine)
    else:
        read_in_file()
        time_entry_check(redmine, 0)
        issue_check(redmine)

#   for issue_data in g_issue_list:
#       print("---- issue[%d] ----" % (issue_data.id))
#       for offset in range(-6, 2):
#           test_print_old_issue_status(issue_data, get_weeks_monday(g_base_day, offset))

    output_datas()

    log_end(start_time)
    return


if __name__ == "__main__":
    main()
