import os
import sys
import re
import time
import datetime
import openpyxl
import pathlib
import shutil
from zipfile import BadZipFile

g_server_sync_ver = '1.0.0'
g_server_path     = ''
g_tgt_dir         = ''
g_out_path        = ''
g_out_file        = ''
g_filter_ext      =  [r'exe', r'bin']
g_filter_dir      =  [r'old', r'backup', r'.git', r'.svn']
g_filter_name     =  [r'^~.*\.(xlsx|xlsm|docx)', r'^.* - コピー\.(.+)$', r'.*Thumbs\.db$']
g_log_file        = None
g_opt_backup      = 0

g_filterd_list    = []
g_deleted_list    = []


#/* ファイル更新に関する設定 */
g_new_file_interactive       = 0           #/* 新規ファイルのコピーについて           （0：確認なし              1：プロンプトで確認）       */
g_new_file_action            = 1           #/* 新規ファイルのデフォルト               （0：ダウンロードしない    1：ダウンロードする)        */
g_update_file_interactive    = 0           #/* サーバー側更新ファイルのコピーについて （0：確認なし              1：プロンプトで確認）       */
g_update_file_action         = 1           #/* サーバー側更新ファイルのデフォルト     （0：更新しない            1：更新            )        */
g_conflict_file_interactive  = 1           #/* ローカル側更新ファイルのコピーについて （0：確認無し              1：プロンプトで確認）       */
g_conflict_file_action       = 0           #/* ローカル側更新ファイルのデフォルト     （0：更新しない            1：更新            ）       */
g_deleted_file_interactive   = 1           #/* サーバー側削除ファイルの削除について   （0：確認なし              1：プロンプトで確認）       */
g_deleted_file_action        = 1           #/* サーバー側削除ファイルのデフォルト     （0：維持                  1：削除            ）       */
g_deleted_local_interactive  = 1           #/* ローカル側削除ファイルの復活について   （0：確認なし              1：プロンプトで確認）       */
g_deleted_local_action       = 1           #/* ローカル側削除ファイルのデフォルト     （0：維持(削除したまま)    1：復活            ）       */


g_target_list  =  []
g_local_list   =  []
g_sync_dt      =  None
g_file_list    =  None

PRE_COL_OFFSET = 10


FILE_NEW            = 0          #/* サーバーにしかないファイル                   */
FILE_SAME           = 1          #/* サーバーと一致しているファイル               */
FILE_SERVER_UPDATED = 2          #/* サーバー上で更新のあったファイル             */
FILE_LOCAL_UPDATED  = 3          #/* ローカルで更新のあったファイル               */
FILE_CONFLICTED     = 4          #/* サーバーとローカル双方で更新のあったファイル */
FILE_LOCAL_ONLY     = 5          #/* ローカルにのみ存在するファイル               */
FILE_SERVER_DELETED = 6          #/* サーバーから無くなったファイル               */
FILE_SERVER_SKIPPED = 7          #/* 明示的に取得しなかったファイル               */
FILE_LOCAL_DELETED  = 8          #/* ローカルで削除したファイル                   */

COL_FILE_NAME       = 3
COL_REL_PATH        = 13
COL_WIDTH_BASE      = 14

SIZE_RJUST          = 12

INPUT_NO            = 0
INPUT_YES           = 1
INPUT_YES_ALL       = 2
INPUT_NO_ALL        = 3

re_time_stamp       = re.compile(r'([0-9]+_[0-9]+)')


#/*****************************************************************************/
#/* ファイルの更新判定情報                                                    */
#/*****************************************************************************/
class cDateTimeSize:
    __slots__ = ['__date', '__time', '__size', '__is_valid']

    def __init__(self):
        self.__date     = '--------'
        self.__time     = '------'
        self.__size     = '-'
        self.__is_valid = False
        return

    #/* タイムスタンプとサイズ情報登録                                        */
    def set_date_time_size(self, date, time, size):
        if (type(date) is datetime.datetime) and (type(time) is datetime.time):
            self.__date     = date.date()
            self.__time     = time.replace(microsecond = 0)
            self.__size     = size
            self.__is_valid = True
        return

    #/* 日付の取得                                                            */
    def get_date(self):
        return self.__date

    #/* 時刻の取得                                                            */
    def get_time(self):
        return self.__time

    #/* サイズの取得                                                          */
    def get_size(self):
        return self.__size

    #/* 有効な情報か？                                                        */
    def is_valid(self):
        return self.__is_valid

    #/* 比較                                                                  */
    def is_equal(self, another_dts):
        if (self.__is_valid) and (another_dts.is_valid()):
            return (self.__date == another_dts.get_date()) and (self.__time == another_dts.get_time()) and (self.__size == another_dts.get_size())

        return False

    #/* テキスト表示情報                                                      */
    def get_print_str(self):
        return f'[{self.__date} {self.__time} {str(self.__size).rjust(SIZE_RJUST)}]'

    #/* タイムスタンプ生成                                                    */
    def get_time_stamp(self):
        date_time_text = str(self.__date).replace('-', '') + '_' + str(self.__time).replace(':', '')
        result = re_time_stamp.match(date_time_text)
        if not (result):
            return '-------- ------            -'

        return result.group(1)


NONE_DTS = cDateTimeSize()


#/*****************************************************************************/
#/* 対象ファイル情報クラス                                                    */
#/*****************************************************************************/
class cFileItem:
    __slots__ = ['file_name', 'rel_path', 'server_dts', 'server_dts_pre', 'local_dts', 'local_dts_pre', 'base_dts', 'base_dts_pre', 'file_attribute', 'server_file', 'local_file']

    def __init__(self, file_name, rel_path):
        self.file_name         = file_name
        self.rel_path          = rel_path

        #/* サーバーのファイル情報 */
        self.server_dts        = cDateTimeSize()
#       self.server_dts_pre    = cDateTimeSize()

        #/* ローカルコピーのファイル情報 */
        self.local_dts         = cDateTimeSize()
#       self.local_dts_pre     = cDateTimeSize()

        #/* サーバーから取得した際のファイル情報 */
        self.base_dts          = cDateTimeSize()
#       self.base_dts_pre      = cDateTimeSize()

        self.file_attribute    = FILE_NEW
        self.server_file       = None
        self.local_file        = None
        return

    def get_server_time_stamp(self):
        return self.server_dts.get_time_stamp()

    def get_local_time_stamp(self):
        return self.local_dts.get_time_stamp()

    def get_base_time_stamp(self):
        return self.base_dts.get_time_stamp()

    def set_server_file(self, cfile):
        self.server_file = cfile
        return

    def set_local_file(self, cfile):
        self.local_file  = cfile
        return

    def set_server_update(self, dts):
#       print(f'set_server_update : {dts.get_print_str()}')
        self.server_dts = dts
        return

    def set_local_update(self, dts):
        self.local_dts = dts
        return

    def set_base_update(self, dts):
        self.base_dts  = dts
        return

    #/* ファイル属性判定 */
    def judge_attribute(self):
        if (self.server_file) and (self.local_file):
            #/* サーバーとローカルにファイルが存在 */
            now_server_dts    = self.server_file.get_dts()
            now_local_dts     = self.local_file.get_dts()
            is_synced         = now_server_dts.is_equal(now_local_dts)

            if (self.base_dts.is_valid()):
                #/* ベースの情報が管理表に存在 */
                is_server_updated = not now_server_dts.is_equal(self.base_dts)
                is_local_updated  = not now_local_dts.is_equal(self.base_dts)

                if (is_synced):
                    #/* サーバーとローカルが一致 */
                    self.file_attribute = FILE_SAME
                elif (is_server_updated) and (is_local_updated):
                    #/* サーバーとローカル両方更新 */
                    self.file_attribute = FILE_CONFLICTED
                elif (is_server_updated):
                    #/* サーバーのみ更新 */
                    self.file_attribute = FILE_SERVER_UPDATED
                elif (is_local_updated):
                    #/* ローカルのみ更新 */
                    self.file_attribute = FILE_LOCAL_UPDATED
                else:
                    print_log(f'Something Wrong. {self.file_name} about Update!')
                    print_log(f'is_server_updated : {is_server_updated},  is_local_updated : {is_local_updated}')
                    print_log(f'server : {self.server_file.get_dts().get_print_str()}')
                    print_log(f'local  : {self.local_file.get_dts().get_print_str()}')
                    print_log(f'base   : {self.base_dts.get_print_str()}')
                    exit(-1)

            else:
                #/* ベースの情報が管理表に無い場合は、単純にファイルのDTS比較で判定する */
                if (is_synced):
                    self.file_attribute = FILE_SAME
                else:
                    self.file_attribute = FILE_CONFLICTED
                pass

            pass

        elif (self.server_file):
            #/* サーバーのみにファイルが存在 */
            now_server_dts    = self.server_file.get_dts()
            if (self.base_dts.is_valid()):
                #/* 管理表上にベースの情報があった場合は、ローカル削除したファイル */
                self.file_attribute = FILE_LOCAL_DELETED

            elif (self.server_dts.is_valid()):
                #/* 管理表上にサーバーファイルの情報があった場合は、意図的に取得しなかったファイル */
                self.file_attribute = FILE_SERVER_SKIPPED

            else:
                #/* 管理表上にサーバーファイルの情報がない場合は、サーバーに新規追加されたファイル */
                self.file_attribute = FILE_NEW

        elif (self.local_file):
            #/* ローカルのみにファイルが存在 */
            if (self.base_dts.is_valid()) or (self.server_dts.is_valid()):
                #/* 管理表上にサーバーまたはベースの情報があった場合は、サーバー側で削除されたファイル */
                self.file_attribute = FILE_SERVER_DELETED

            else:
                #/* 管理表上にサーバーファイルの情報がない場合は、ローカルで作成されたファイル */
                self.file_attribute = FILE_LOCAL_ONLY

        else:
            #/* 管理表のみにファイルが存在 */
            self.file_attribute = FILE_SERVER_DELETED

#       print(f'[{self.get_attribute_text().ljust(10)}] : {self.rel_path}{self.file_name}')
        return

    #/* 全DTSの表示 */
    def print_all(self):
        server_dts_txt = self.server_dts.get_print_str()
        local_dts_txt  = self.local_dts.get_print_str()
        base_dts_txt   = self.base_dts.get_print_str()
        if (self.server_file) and (self.local_file):
            print_log(f'cFileItem1 : {server_dts_txt} {local_dts_txt} {base_dts_txt} : [{self.file_name}]')
        elif (self.server_file):
            print_log(f'cFileItem2 : {server_dts_txt} {local_dts_txt} {base_dts_txt} : [{self.file_name}]')
        elif (self.local_file):
            print_log(f'cFileItem3 : {server_dts_txt} {local_dts_txt} {base_dts_txt} : [{self.file_name}]')
        else:
            print_log(f'cFileItem4 : {server_dts_txt} {local_dts_txt} {base_dts_txt} : [{self.file_name}]')

    #/* 管理ファイルからの1行読み出し */
    def read_ws_row(self, ws, row):
        col = 2
        attribute = ws.cell(row, col).value
        col += 1

#       #/* ファイル名は読み飛ばす */
#       self.file_name       = ws.cell(row, col).value
        col += 1

        #/* サーバー側ファイル情報 */
        server_date          = ws.cell(row, col).value
        col += 1
        server_time          = ws.cell(row, col).value
        col += 1
        server_size          = ws.cell(row, col).value
        col += 1

        server_dts           = cDateTimeSize()
        server_dts.set_date_time_size(server_date, server_time, server_size)
        self.set_server_update(server_dts)

        #/* ローカル側ファイル情報 */
        local_date    = ws.cell(row, col).value
        col += 1
        local_time    = ws.cell(row, col).value
        col += 1
        local_size    = ws.cell(row, col).value
        col += 1

        local_dts     = cDateTimeSize()
        local_dts.set_date_time_size(local_date, local_time, local_size)
        self.set_local_update(local_dts)

        #/* ファイル取得時情報 */
        base_date     = ws.cell(row, col).value
        col += 1
        base_time     = ws.cell(row, col).value
        col += 1
        base_size     = ws.cell(row, col).value
        col += 1

        base_dts      = cDateTimeSize()
        base_dts.set_date_time_size(base_date, base_time, base_size)
        self.set_base_update(base_dts)

#       #/* 相対パスは読み飛ばす */
#       self.rel_path        = ws.cell(row, col).value
        col += 1
        return

    #/* 管理ファイルへの1行書き込み */
    def write_ws_row(self, ws, row):
        col  = 2

        #/* 同期状態 */
        ws.cell(row, col).value = self.get_attribute_text()
        col += 1

        #/* ファイル名 */
        ws.cell(row, col).value = self.file_name
        col += 1

        #/* サーバーファイル情報 */
        if (self.server_file):
            server_dts = self.server_file.get_dts()
        else:
            server_dts = NONE_DTS

        ws.cell(row, col).value = server_dts.get_date()
        col += 1
        ws.cell(row, col).value = server_dts.get_time()
        col += 1
        ws.cell(row, col).value = server_dts.get_size()
        col += 1

        #/* ローカルファイル情報 */
        if (self.local_file):
            local_dts = self.local_file.get_dts()
        else:
            local_dts = NONE_DTS

        ws.cell(row, col).value = local_dts.get_date()
        col += 1
        ws.cell(row, col).value = local_dts.get_time()
        col += 1
        ws.cell(row, col).value = local_dts.get_size()
        col += 1

        #/* ベースファイル情報 */
        ws.cell(row, col).value = self.base_dts.get_date()
        col += 1
        ws.cell(row, col).value = self.base_dts.get_time()
        col += 1
        ws.cell(row, col).value = self.base_dts.get_size()
        col += 1

        #/* 相対パス */
        ws.cell(row, col).value = self.rel_path
        col += 1
        return

    def get_attribute_text(self):
        text = '-'
        if (self.file_attribute == FILE_NEW):
            text = 'NEW'
        elif (self.file_attribute == FILE_SAME):
            text = 'SAME'
        elif (self.file_attribute == FILE_SERVER_UPDATED):
            text = 'UPDATED'
        elif (self.file_attribute == FILE_LOCAL_UPDATED):
            text = 'LOCAL'
        elif (self.file_attribute == FILE_CONFLICTED):
            text = 'CONFLICTED'
        elif (self.file_attribute == FILE_LOCAL_ONLY):
            text = 'LOCAL_ONLY'
        elif (self.file_attribute == FILE_SERVER_DELETED):
            text = 'DELETED'
        elif (self.file_attribute == FILE_SERVER_SKIPPED):
            text = 'SKIPPED'
        elif (self.file_attribute == FILE_LOCAL_DELETED):
            text = 'MISSING'

        return text


#/*****************************************************************************/
#/* ファイルリスト情報クラス                                                  */
#/*****************************************************************************/
class cFileItemList:
    def __init__(self):
        self.target_path   = '-'
        self.sync_date     = None
        self.sync_time     = None
        self.sync_ver      = ''
        self.log_file      = '-'
        self.pre_sync_date = None
        self.pre_sync_time = None
        self.pre_sync_ver  = ''
        self.pre_log_file  = '-'
        self.items         = []        #/* cFileItem */
        return

    #/* 管理ファイルからのファイルの情報読み込み */
    def read_worksheet(self, ws):
        row = 2
        col = 3
        self.target_path   = ws.cell(row, col).value
        row += 1
        self.sync_date     = ws.cell(row, col).value
        self.pre_sync_date = ws.cell(row, col + PRE_COL_OFFSET).value
        row += 1
        self.sync_time     = ws.cell(row, col).value
        self.pre_sync_time = ws.cell(row, col + PRE_COL_OFFSET).value
        row += 1

        self.sync_ver      = ws.cell(row, col).value
        self.pre_sync_ver  = ws.cell(row, col + PRE_COL_OFFSET).value
        row += 1

        self.log_file      = ws.cell(row, col).value
        self.pre_log_file  = ws.cell(row, col + PRE_COL_OFFSET).value

        row += 3
        col  = 3
        while (ws.cell(row, col).value != None):
            item = cFileItem(ws.cell(row, COL_FILE_NAME).value, ws.cell(row, COL_REL_PATH).value)
            item.read_ws_row(ws, row)
            self.items.append(item)
#           item.print_all()
            row += 1
        return

    #/* サーバーに実在するファイルの情報追加 */
    def add_server_file(self, cfile):
        for item in self.items:
            if (cfile.file_name == item.file_name) and (cfile.rel_path == item.rel_path):
                #/* サーバーファイルの更新情報を更新する */
                item.set_server_file(cfile)
                return

        #/* 見つからなかった場合は、itemsに足す */
        item = cFileItem(cfile.file_name, cfile.rel_path)
        item.set_server_file(cfile)
        self.items.append(item)
        return

    #/* ローカルに実在するファイルの情報追加 */
    def add_local_file(self, cfile):
        for item in self.items:
            if (cfile.file_name == item.file_name) and (cfile.rel_path == item.rel_path):
                item.set_local_file(cfile)
                return

        #/* ローカルにのみ存在するファイルも、一応itemsに足す */
        item = cFileItem(cfile.file_name, cfile.rel_path)
        item.set_local_file(cfile)
        self.items.append(item)
        return


#/*****************************************************************************/
#/* ファイル情報クラス                                                        */
#/*****************************************************************************/
class cFileInfo:
    __slots__ = ['file_name', 'rel_path', 'abs_path', '__date_time_size']

    def __init__(self, nw_path, root_path):
        self.file_name      = nw_path.name
#       self.rel_path       = str(nw_path).replace(self.file_name, '')
        self.rel_path       = str(nw_path).rstrip(self.file_name)
        self.rel_path       = self.rel_path.replace(root_path, '', 1)
        self.abs_path       = nw_path

        dts                 = cDateTimeSize()
        stat_info           = nw_path.stat()
        dt                  = datetime.datetime.fromtimestamp(stat_info.st_mtime)
        dts.set_date_time_size(dt, dt.time(), stat_info.st_size)
        self.__date_time_size = dts
#       print_log(f'cFileInfo : [{dts.date} {dts.time} {str(dts.size).rjust(SIZE_RJUST)}] {nw_path}')
        return

    def get_time_stamp(self):
        return self.__date_time_size.get_time_stamp()

    def get_dts(self):
        return self.__date_time_size


#/*****************************************************************************/
#/* ログ出力                                                                  */
#/*****************************************************************************/
def print_log(text):
    global g_log_file
    print(text)
    print(text, file=g_log_file)


#/*****************************************************************************/
#/* 処理開始ログ                                                              */
#/*****************************************************************************/
def log_start():
    global g_log_file
    global g_out_path
    global g_tgt_dir
    global g_sync_dt

    g_sync_dt = datetime.datetime.now()
    time_stamp = g_sync_dt.strftime('%Y%m%d_%H%M%S')
    date_stamp = g_sync_dt.strftime('%Y%m%d')

    log_path   = g_tgt_dir + '\\.SrvSync\\ServerSync_' + time_stamp + '.txt'
    g_log_file = open(log_path, "w")
#   sys.stdout = log_file

    start_time = time.perf_counter()
    g_sync_dt = datetime.datetime.now()
    print_log("処理開始 : " + str(g_sync_dt))
    print_log("----------------------------------------------------------------------------------------------------------------")
    return start_time


#/*****************************************************************************/
#/* 処理終了ログ                                                              */
#/*****************************************************************************/
def log_end(start_time):
    end_time = time.perf_counter()
    now = datetime.datetime.now()
    print_log("----------------------------------------------------------------------------------------------------------------")
    print_log("処理終了 : " + str(now))
    second = int(end_time - start_time)
    msec   = ((end_time - start_time) - second) * 1000
    minute = second / 60
    second = second % 60
    print_log("  %dmin %dsec %dmsec" % (minute, second, msec))
    return


#/*****************************************************************************/
#/* サブディレクトリの生成                                                    */
#/*****************************************************************************/
def make_directory(dirname):
#   print_log("make dir! %s" % dirname)
    os.makedirs(os.path.join(dirname), exist_ok = True)


#/*****************************************************************************/
#/* コマンドライン引数処理                                                    */
#/*****************************************************************************/
def check_command_line_option():
    global g_server_path
    global g_tgt_dir
    global g_out_path
    global g_out_file
    global g_opt_backup

    sys.argv.pop(0)
    for arg in sys.argv:
        if (arg == '-backup'):
            g_opt_backup = 1
        elif (os.path.isdir(arg)):
            g_server_path = arg
        else:
            print('invarid arg : %s' & arg)

    if (g_server_path == ''):
        print('server_sync.py [target_path]')
        exit(-1)

    g_server_path = g_server_path.rstrip('\\')                             #/* 末尾のバックスラッシュを削除 */
    g_tgt_dir  = pathlib.WindowsPath(g_server_path).name
    g_out_path = g_tgt_dir + '\\.SrvSync'
    g_out_file = g_out_path + '\\ServerSync.xlsx'
    return


#/*****************************************************************************/
#/* ディレクトリ名のフィルタ                                                  */
#/*****************************************************************************/
def dir_filter_check(file_path):
    global g_filter_dir

    if (file_path.name == '.SrvSync'):
        return True

    if (file_path.is_dir()):
        for filter in g_filter_dir:
            if (filter == file_path.name):
                return True

    return False


#/*****************************************************************************/
#/* 拡張子のフィルタ                                                          */
#/*****************************************************************************/
def extension_filter_check(file_path):
    global g_filter_ext

#   print(f'file_path:{file_path},  suffix:{file_path.suffix}')
    if (file_path.is_file()):
        for filter in g_filter_ext:
            if ('.' + filter == file_path.suffix):
                return True

    return False


#/*****************************************************************************/
#/* ファイル名のmatchフィルタ                                                 */
#/*****************************************************************************/
def match_filter_check(file_path):
    global g_filter_name

    if (file_path.is_file()):
        for filter in g_filter_name:
            if (result := re.match(filter, str(file_path))):
                return True

    return False


#/*****************************************************************************/
#/* パス検索                                                                  */
#/*****************************************************************************/
def search_target_path(level, path, list, root):
    print_log('Search     : %s' % (path))
    network_path = pathlib.WindowsPath(path)
    if (network_path.exists()):
        files = network_path.iterdir()
        for file_name in files:
#           print_log('file_name : %s' % (file_name))
            if (dir_filter_check(file_name)):
#               print_log('Filter Directory! [%s]' % (file_name))
                continue
            elif (extension_filter_check(file_name)):
#               print_log('Filter Extension! [%s]' % (file_name))
                continue
            elif (match_filter_check(file_name)):
#               print_log('Filter match! [%s]' % (file_name))
                continue

            if (file_name.is_dir()):
                search_target_path(level + 1, file_name, list, root)
            else:
                try:
                    file_info = cFileInfo(file_name, root)
                except FileNotFoundError:
                    if len(str(file_name)) > 260:
                        print(f'ファイル[{file_name}]のパスが長すぎます')
                    else:
                        print(f'ファイル[{file_name}]が見つかりません')
                else:
                    list.append(file_info)

    else:
        print_log("network path is not available!")

    return


#/*****************************************************************************/
#/* ファイルリスト入力                                                        */
#/*****************************************************************************/
def in_file_list():
    global g_out_file
    global g_opt_backup
    global g_file_list

    if (os.path.isfile(g_out_file)):
        writable = os.access(g_out_file, os.W_OK)
        if not (writable):
            print_log('[%s]が読み取り専用です' % (g_out_file))
            exit(-1)

    wb = None
    try:
        wb = openpyxl.load_workbook(g_out_file, data_only=True)
    except PermissionError:
        pass
    except FileNotFoundError:
        print(f'ファイル[{g_out_file}]が見つかりません')
    except BadZipFile:
        print(f'ファイル[{g_out_file}]を開くことができません')
        exit(-1)
    else:
        pass

#   print_log(f'g_opt_backup is {g_opt_backup}')
    if (wb != None) and (g_opt_backup == 1):
        path = pathlib.WindowsPath(g_out_file)
        file_info = cFileInfo(path, g_server_path)
        time_stamp = file_info.get_time_stamp()
        back_up_path = g_out_file.replace(r'.xlsx', r'_' + time_stamp + r'.xlsx')
        print_log(f'リストファイルをバックアップします [{g_out_file}] -> [{back_up_path}]')
        wb.save(g_out_file.replace(r'.xlsx', r'_' + time_stamp + r'.xlsx'))

    g_file_list = cFileItemList()
    if (wb != None):
        ws = wb['FileList']
        g_file_list.read_worksheet(ws)

        if (g_file_list.target_path != g_server_path):
            print(f'同期パスが異なります {g_server_path} vs {g_file_list.target_path}')
            exit(-1)

    return


#/*****************************************************************************/
#/* サーバー -> ローカルへのファイルコピー                                    */
#/*****************************************************************************/
def copy_item_file(item):
    global g_tgt_dir

    make_directory(g_tgt_dir + item.rel_path)
    local_path = g_tgt_dir + item.rel_path + item.file_name
#   print_log(f'copy new file {local_path}')

    try:
        shutil.copy2(item.server_file.abs_path,  local_path)
        path = pathlib.WindowsPath(local_path)
        local_file_info = cFileInfo(path, g_tgt_dir)
        item.set_local_file(local_file_info)
        item.set_base_update(local_file_info.get_dts())
    except FileNotFoundError:
        print_log(f"{item.server_file.abs_path}が見つかりませんでした")
    except:
        print_log(f"{item.server_file.abs_path}のコピーに失敗しました")

    return


#/*****************************************************************************/
#/* プロンプトによる意思確認                                                  */
#/*****************************************************************************/
def check_key_input(text):
     while(True):
         input_key = input(text + ' (Y or N)')
         if (input_key.upper() == 'Y'):
             return INPUT_YES
         elif (input_key.upper() == 'N'):
             return INPUT_NO
         elif (input_key.upper() == 'YA'):
             return INPUT_YES_ALL
         elif (input_key.upper() == 'NA'):
             return INPUT_NO_ALL


#/*****************************************************************************/
#/* 更新されたファイルのコピー                                                */
#/*****************************************************************************/
def copy_updated_files():
    global g_target_list
    global g_local_list
    global g_tgt_dir

    global g_new_file_interactive
    global g_new_file_action
    global g_update_file_interactive
    global g_update_file_action
    global g_conflict_file_interactive
    global g_conflict_file_action
    global g_deleted_file_interactive
    global g_deleted_file_action
    global g_deleted_local_interactive
    global g_deleted_local_action

    #/* 先に、ファイルの更新属性を判定する */
    for item in g_file_list.items:
        item.judge_attribute()

    for item in g_file_list.items:
        if (item.server_file):
            server_dts = item.server_file.get_dts().get_print_str()
        else:
            server_dts = NONE_DTS.get_print_str()

        if (item.file_attribute == FILE_NEW) or (item.file_attribute == FILE_SERVER_SKIPPED):
            #/* 新規ファイル */
            action = g_new_file_action
            if (g_new_file_interactive):
                action = check_key_input(f'サーバーの新規ファイル : {item.rel_path} {item.file_name}をダウンロードしますか？')

            if (action == INPUT_YES):
                print_log(f'Added File     : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
            elif (action == INPUT_YES_ALL):
                print_log(f'Add All        : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
                g_new_file_action      = 1
                g_new_file_interactive = 0
            elif (action == INPUT_NO):
                print_log(f'Skip(A)        : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                item.file_attribute = FILE_SERVER_SKIPPED
            elif (action == INPUT_NO_ALL):
                print_log(f'Skip(A) All    : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                item.file_attribute = FILE_SERVER_SKIPPED
                g_new_file_action      = 0
                g_new_file_interactive = 0

        elif (item.file_attribute == FILE_SAME):
            #/* すでに同期しているファイル */
            if not (item.base_dts.is_valid()):
                print_log(f'Already Exists : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                item.set_base_update(item.local_file.get_dts())

        elif (item.file_attribute == FILE_SERVER_UPDATED):
            #/* サーバー側で更新のあったファイル */
            action = g_update_file_action
            if (g_update_file_interactive):
                action = check_key_input(f'サーバーの更新ファイル : {item.rel_path} {item.file_name}をダウンロードしますか？')

            if (action == INPUT_YES):
                print_log(f'Updated File   : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
            elif (action == INPUT_YES_ALL):
                print_log(f'Update All     : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
                g_update_file_action      = 1
                g_update_file_interactive = 0
            elif (action == INPUT_NO):
                print_log(f'Skip(U)        : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
            elif (action == INPUT_NO_ALL):
                print_log(f'Skip(U) All    : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                g_update_file_action      = 0
                g_update_file_interactive = 0

        elif (item.file_attribute == FILE_LOCAL_UPDATED):
            #/* ローカル側で更新のあったファイル */
            pass

        elif (item.file_attribute == FILE_CONFLICTED):
            #/* サーバーとローカル双方で更新のあったファイル */
            action = g_conflict_file_action
            if (g_conflict_file_interactive):
                action = check_key_input(f'！コンフリクトファイル！ : {item.rel_path} {item.file_name}を上書き更新しますか？')
                if (action == INPUT_YES):
                    action = check_key_input(f'ローカル変更ファイルが上書きされますが、間違いないですか？')

            if (action == INPUT_YES):
                print_log(f'Overwrite File : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
            elif (action == INPUT_YES_ALL):
                print_log(f'Overwrite All  : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
                g_conflict_file_action      = 1
                g_conflict_file_interactive = 0
            elif (action == INPUT_NO):
                print_log(f'Skip(C)        : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
            elif (action == INPUT_NO_ALL):
                print_log(f'Skip(C) All    : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                g_conflict_file_action      = 0
                g_conflict_file_interactive = 0

        elif (item.file_attribute == FILE_LOCAL_ONLY):
            #/* ローカルのみに存在するファイル */
            pass

        elif (item.file_attribute == FILE_SERVER_DELETED):
            #/* サーバーで削除されたファイル */
            if (item.local_file):
                local_dts = item.local_file.get_dts().get_print_str()

                action = g_deleted_file_action
                if (g_deleted_file_interactive):
                    action = check_key_input(f'サーバーで削除されたファイル : {item.rel_path} {item.file_name}を削除しますか？')

                if (action == INPUT_YES):
                    print_log(f'Delete         : {local_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                    os.remove(g_tgt_dir + item.rel_path + item.file_name)
                elif (action == INPUT_YES_ALL):
                    print_log(f'Delete All     : {local_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                    os.remove(g_tgt_dir + item.rel_path + item.file_name)
                    g_deleted_file_action = 1
                    g_deleted_file_interactive = 0
                elif (action == INPUT_NO):
                    print_log(f'Skip(D)        : {local_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                elif (action == INPUT_NO_ALL):
                    print_log(f'Skip(D)  All   : {local_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                    g_deleted_file_action = 0
                    g_deleted_file_interactive = 0
            else:
                base_dts = item.base_dts.get_print_str()
                print_log(f'Already Delete : {base_dts} {g_tgt_dir + item.rel_path + item.file_name}')

        elif (item.file_attribute == FILE_LOCAL_DELETED):
            #/* ローカルで削除したファイル */
            action = g_deleted_local_action

            if (g_deleted_local_interactive):
                action = check_key_input(f'ローカルで削除したファイル : {item.rel_path} {item.file_name}を復元しますか？')

            if (action == INPUT_YES):
                print_log(f'Missing File   : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
            elif (action == INPUT_YES_ALL):
                print_log(f'Missing All    : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                copy_item_file(item)
                g_deleted_local_action      = 1
                g_deleted_local_interactive = 0
            elif (action == INPUT_NO):
                print_log(f'Skip(M)        : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
            elif (action == INPUT_NO_ALL):
                print_log(f'Skip(M) All    : {server_dts} {g_tgt_dir + item.rel_path + item.file_name}')
                g_deleted_local_action      = 0
                g_deleted_local_interactive = 0


    return


#/*****************************************************************************/
#/* ファイルリスト出力                                                        */
#/*****************************************************************************/
def out_file_list():
    global g_sync_dt
    global g_target_list
    global g_out_file
    global g_file_list
    global g_server_sync_ver

    time_stamp = g_sync_dt.strftime('%Y%m%d_%H%M%S')
    log_path = g_out_path + '\\ServerSync_' + time_stamp + '.txt'

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = 'FileList'

    row = 2
    col = 2
    ws.cell(row, col                     ).value = '対象パス'
    ws.cell(row, col + 1                 ).value = g_server_path
    row += 1
    ws.cell(row, col                     ).value = '同期 日付'
    ws.cell(row, col + 1                 ).value = g_sync_dt.date()
    ws.cell(row, col + PRE_COL_OFFSET    ).value = '前回同期 日付'
    if (g_file_list.sync_date):
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = g_file_list.sync_date.date()
    else:
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = '-'
    row += 1
    ws.cell(row, col                     ).value = '同期 時間'
    ws.cell(row, col + 1                 ).value = g_sync_dt.time()
    ws.cell(row, col + PRE_COL_OFFSET    ).value = '前回同期 時間'
    if (g_file_list.sync_time):
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = g_file_list.sync_time
    else:
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = '-'
    row += 1

    ws.cell(row, col                     ).value = 'バージョン'
    ws.cell(row, col + 1                 ).value = g_server_sync_ver
    ws.cell(row, col + PRE_COL_OFFSET    ).value = '前回バージョン'
    if (g_file_list.sync_ver):
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = g_file_list.sync_ver
    else:
        ws.cell(row, col + PRE_COL_OFFSET + 1).value = '-'
    row += 1

    ws.cell(row, col                     ).value = 'ログ'
    ws.cell(row, col + 1                 ).value = log_path
    ws.cell(row, col + PRE_COL_OFFSET    ).value = '前回ログ'
    ws.cell(row, col + PRE_COL_OFFSET + 1).value = g_file_list.log_file
    ws.column_dimensions['B'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['C'].width              =  60
    ws.column_dimensions['D'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['E'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['F'].width              =  COL_WIDTH_BASE + 2
    ws.column_dimensions['G'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['H'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['I'].width              =  COL_WIDTH_BASE + 2
    ws.column_dimensions['J'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['K'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['L'].width              =  COL_WIDTH_BASE
    ws.column_dimensions['M'].width              =  COL_WIDTH_BASE

    row += 2
    col = 2
    ws.cell(row, col).value = '更新'
    col += 1
    ws.cell(row, col).value = 'ファイル名'
    col += 1
    ws.cell(row, col).value = 'サーバー日付'
    col += 1
    ws.cell(row, col).value = 'サーバー時間'
    col += 1
    ws.cell(row, col).value = 'サーバーサイズ'
    col += 1
    ws.cell(row, col).value = 'ローカル日付'
    col += 1
    ws.cell(row, col).value = 'ローカル時間'
    col += 1
    ws.cell(row, col).value = 'ローカルサイズ'
    col += 1
    ws.cell(row, col).value = '取得時日付'
    col += 1
    ws.cell(row, col).value = '取得時時間'
    col += 1
    ws.cell(row, col).value = '取得時サイズ'
    col += 1
    ws.cell(row, col).value = '相対パス'
    col += 1

    row += 1
    sorted_items = sorted(g_file_list.items, key=lambda item: item.rel_path)

    for item in sorted_items:
        if (item.server_file):
#           print_log('[%s] : %s' % (item.rel_path, item.file_name))
            item.write_ws_row(ws, row)
            row += 1
        elif (item.local_file):
            if (item.file_attribute == FILE_SERVER_DELETED):
                print_log('サーバーから削除されたファイル[%s] : %s' % (item.rel_path, item.file_name))
                item.write_ws_row(ws, row)
                row += 1
            else:
                print_log('ローカルオンリーのファイル[%s] : %s' % (item.rel_path, item.file_name))
        else:
#           print_log('管理表にしかないファイル[%s] : %s' % (item.rel_path, item.file_name))
            item.write_ws_row(ws, row)
            row += 1

    while True:
        try:
            wb.save(g_out_file)
        except PermissionError:
            input(f"ServerSync.xlsxが更新できません。クローズしてから、Enterを入力してください。")
        else:
            print(f'Excel file saved successfully at {g_out_file}')
            break
    return


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    global g_out_path
    global g_target_list
    global g_server_path
    global g_local_list
    global g_tgt_dir
    global g_file_list

    check_command_line_option()
    make_directory(g_out_path)
    start_time = log_start()

    print_log('-------------------------------------------- Read Previous Results ---------------------------------------------')
    in_file_list()

    print_log('-------------------------------------------- List up Server Files ----------------------------------------------')
    search_target_path(0, g_server_path, g_target_list, g_server_path)

    print_log('-------------------------------------------- add_server_file      ----------------------------------------------')
    for target in g_target_list:
        g_file_list.add_server_file(target)

    print_log('-------------------------------------------- List up Local Files  ----------------------------------------------')
    search_target_path(0, g_tgt_dir,     g_local_list,  g_tgt_dir)

    print_log('-------------------------------------------- add_local_file       ----------------------------------------------')
    for local in g_local_list:
        g_file_list.add_local_file(local)

    print_log('-------------------------------------------- Copy Updated Files   ----------------------------------------------')
    copy_updated_files()

    print_log('-------------------------------------------- Write Results        ----------------------------------------------')
    out_file_list()

    log_end(start_time)
    return


if __name__ == "__main__":
    main()





