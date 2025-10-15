import os
import sys
import re
import time
import datetime
import openpyxl
import win32com.client
from openpyxl.drawing.image import Image as XLImage
from pdf2image import convert_from_path


DEFAULT_START_PAGE = 1
DEFAULT_END_PAGE   = 10000

ROW_OFFSET  = 3
COL_OFFSET  = 2
ROW_PADDING = 2


TMP_FOLDER    = 'doc_fulfill_tmp'

g_target_file = ''
g_dpi         = 100
g_start_page  = DEFAULT_START_PAGE
g_end_page    = DEFAULT_END_PAGE
g_template    = ''
g_output      = 'output.xlsx'
g_on_memory   = True


RE_PAGE_OPTION = re.compile(r'([0-9]*):([0-9]*)')


#/*****************************************************************************/
#/* コマンドライン引数処理                                                    */
#/*****************************************************************************/
def check_command_line_option():
    global g_target_file
    global g_template
    global g_dpi
    global g_start_page
    global g_end_page
    global g_output
    global g_on_memory

    sys.argv.pop(0)
    while(len(sys.argv)):
        arg = sys.argv.pop(0)
        if (os.path.isfile(arg)):
            if (arg.lower().endswith('.pdf')) or (arg.lower().endswith('.docx')):
                g_target_file = arg
            else:
                print(f'PDFまたはdocxファイルを指定してください')
                exit(-1)
        elif (arg == '-dpi'):
            try:
                g_dpi = int(sys.argv.pop(0))
            except Exception as e:
                print('DPIを数値で指定してください')
                exit(-1)
        elif (arg == '-p'):
            try:
                pages = sys.argv.pop(0)
#               print(f'pages : {pages}')
                if (result := RE_PAGE_OPTION.match(pages)):
                    g_start_page = int(result.group(1) or DEFAULT_START_PAGE)
                    g_end_page   = int(result.group(2) or DEFAULT_END_PAGE)
                    if (g_start_page > g_end_page):
                        print(f'PAGEの指定が不正です ({start_page} > {end_page})')
                        exit(-1)
                else:
                    print('PAGEを数値で指定してください')
                    exit(-1)
            except Exception as e:
                print('PAGEを数値で指定してください')
                exit(-1)
        elif (arg == '-t'):
            try:
                g_template = sys.argv.pop(0)
                if (not os.path.isfile(g_template)) or (not g_template.lower().endswith('.xlsx')):
                    print(f'テンプレートファイルを指定してください')
                    exit(-1)

            except Exception as e:
                print('テンプレートファイルを指定してください')
                exit(-1)
        elif (arg == '-o'):
            try:
                g_output = sys.argv.pop(0)
                if (not g_output.lower().endswith('.xlsx')):
                    print(f'出力ファイル名を指定してください')
                    exit(-1)

            except Exception as e:
                print(f'出力ファイル名を指定してください')
                exit(-1)
        elif (arg == '-png'):
            g_on_memory = False
        else:
            print("invalid arg : %s" % arg)

    if (g_target_file == ''):
        print(f'PDFファイルを指定してください')
        exit(-1)

    return



#/*****************************************************************************/
#/* 処理開始ログ                                                              */
#/*****************************************************************************/
def log_start():

    now = datetime.datetime.now()

    time_stamp = now.strftime('%Y%m%d_%H%M%S')
    log_path = 'doc_fulfill_' + time_stamp + '.txt'
    log_file = open(log_path, "w")
    sys.stdout = log_file

    start_time = time.perf_counter()
    now = datetime.datetime.now()
    print("処理開始 : " + str(now))
    print ("----------------------------------------------------------------------------------------------------------------")
    return start_time


#/*****************************************************************************/
#/* 処理終了ログ                                                              */
#/*****************************************************************************/
def log_end(start_time):
    end_time = time.perf_counter()
    now = datetime.datetime.now()
    print ("----------------------------------------------------------------------------------------------------------------")
    print("処理終了 : " + str(now))
    second = int(end_time - start_time)
    msec   = ((end_time - start_time) - second) * 1000
    minute = second / 60
    second = second % 60
    print("  %dmin %dsec %dmsec" % (minute, second, msec))
    return



#/*****************************************************************************/
#/* DOCXからPDFへの変換                                                       */
#/*****************************************************************************/
def convert_docx_to_pdf(target_docx):
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False

    doc = word.Documents.Open(target_docx)

    tmp_pdf = os.path.abspath(f'{TMP_FOLDER}\\{os.path.basename(target_docx)}.pdf')
    doc.SaveAs(tmp_pdf, FileFormat=17)  # 17 = PDF
    doc.Close()
    word.Quit()
    return tmp_pdf


#/*****************************************************************************/
#/* PDFから画像への変換                                                       */
#/*****************************************************************************/
def convert_pdf_to_image(target_pdf):
    images = []
    pages = convert_from_path(target_pdf, dpi=g_dpi, first_page = g_start_page, last_page = g_end_page)
    for i, page in enumerate(pages, start=1):
        img = XLImage(page)  # ← ファイルに保存しなくてもOK
        print(f'img [{img.width} x {img.height}]')
        images.append(img)

        if not g_on_memory:
            img_path = f'{TMP_FOLDER}\\{os.path.basename(target_pdf)}_p{g_start_page + i - 1}.png'
            page.save(img_path, "PNG")
    return images



#/*****************************************************************************/
#/* Excelへの貼り付け                                                         */
#/*****************************************************************************/
def output_excel(target_xlsx, images):
    if (g_template != ''):
        wb = openpyxl.load_workbook(g_template, data_only=True)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    row = ROW_OFFSET
    col = openpyxl.utils.get_column_letter(COL_OFFSET)
    page = g_start_page
    for img in images:
        ws.cell(row, 1).value = f'Page {page}'
        row += 1
        ws.add_image(img, f"{col}{row}")
        row_height = int(img.height / 25) + 1
        row += (row_height + ROW_PADDING)
        page += 1

    wb.save(g_output)
    return


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    global g_target_file

    check_command_line_option()
    start_time = log_start()

    if not g_on_memory:
        os.makedirs(TMP_FOLDER, exist_ok = True)

    if g_target_file.lower().endswith('.docx'):
        os.makedirs(TMP_FOLDER, exist_ok = True)
        target_abs_path = os.path.abspath(g_target_file)
        g_target_file = convert_docx_to_pdf(target_abs_path)

    images = convert_pdf_to_image(g_target_file)
    output_excel(g_output, images)
    log_end(start_time)
    return


if __name__ == "__main__":
    main()
